from io import BytesIO
from collections import defaultdict

from django.contrib import messages
from django.db import models, transaction
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from accounts.decorators import role_required
from accounts.models import User
from core.student_ordering import (
    order_queryset_by_student_name,
    resolve_student_order,
    student_order_context,
)
from enrollment.models import Enrollment

from .forms import CompetencyForm, CourseForm, GradeForm, GradeRecordForm, IndicatorForm, SectionForm
from .models import (
    AcademicYear,
    Competency,
    Course,
    Grade,
    GradeRecord,
    Indicator,
    IndicatorGrade,
    Period,
    Section,
    Unit,
    TeacherCourseAssignment,
    calculate_mode_grade,
    GradeSubmissionLock,
)
from .sync import sync_teacher_course_assignments


POLY_COURSE_NAMES = {
    'ingles',
    'computacion',
    'arte',
    'educacion fisica',
    'robotica',
    'electronica',
}


def _normalize_course_name(name):
    replacements = str.maketrans({
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'a', 'É': 'e', 'Í': 'i', 'Ó': 'o', 'Ú': 'u',
    })
    return (name or '').translate(replacements).strip().lower()


def _is_poly_assignment(assignment):
    if not assignment:
        return False
    return (
        assignment.course.is_poly_course
        or _normalize_course_name(assignment.course.name) in POLY_COURSE_NAMES
    )


def _assignment_classroom_label(assignment):
    if assignment.section:
        return f"{assignment.section.grade.name} {assignment.section.name}"
    if assignment.grade:
        return assignment.grade.name
    return "Sin seccion"


def _assignment_select_label(assignment):
    if _is_poly_assignment(assignment):
        return f"{assignment.course.name} - {_assignment_classroom_label(assignment)}"
    return assignment.course.name


def _competency_source_assignment(assignment):
    if not _is_poly_assignment(assignment):
        return assignment

    sibling_assignments = TeacherCourseAssignment.objects.filter(
        teacher=assignment.teacher,
        course=assignment.course,
        academic_year=assignment.academic_year,
    )
    source = (
        sibling_assignments
        .filter(competencies__isnull=False)
        .order_by('section__grade__name', 'section__name', 'id')
        .first()
    )
    return source or assignment


@role_required('admin', 'director')
def manage_grade_locks(request):
    """View for Director to see and toggle submission locks."""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, "No hay un año académico activo.")
        return redirect('academic_dashboard')

    assignments = TeacherCourseAssignment.objects.filter(academic_year=active_year).select_related('teacher', 'course', 'section__grade')
    periods = Period.objects.filter(academic_year=active_year).order_by('name')
    
    # Pre-build a lookup dictionary for locks: (teacher_id, course_id, section_id, period_id) -> is_locked
    locks_qs = GradeSubmissionLock.objects.filter(period__academic_year=active_year)
    lock_lookup = {
        (l.teacher_id, l.course_id, l.section_id, l.period_id): l.is_locked
        for l in locks_qs
    }

    # Flatten data for the template to avoid nested loops with complex lookups
    lock_data = []
    for assignment in assignments:
        for period in periods:
            key = (assignment.teacher_id, assignment.course_id, assignment.section_id, period.id)
            is_locked = lock_lookup.get(key, False)
            lock_data.append({
                'assignment': assignment,
                'period': period,
                'is_locked': is_locked,
                'key': f"{assignment.teacher_id}_{assignment.course_id}_{assignment.section_id}_{period.id}"
            })
    
    context = {
        'lock_data': lock_data,
    }
    return render(request, 'academic/manage_grade_locks.html', context)


@role_required('admin', 'director')
def toggle_grade_lock(request):
    """Director toggles the lock for a specific submission."""
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        course_id = request.POST.get('course_id')
        section_id = request.POST.get('section_id')
        period_id = request.POST.get('period_id')
        action = request.POST.get('action')  # 'lock' or 'unlock'
        next_url = request.POST.get('next')

        if action not in {'lock', 'unlock'}:
            messages.error(request, "Accion invalida.")
            return redirect(next_url or 'manage_grade_locks')

        section_id = section_id or None

        lock, created = GradeSubmissionLock.objects.update_or_create(
            teacher_id=teacher_id,
            course_id=course_id,
            section_id=section_id,
            period_id=period_id,
            defaults={'is_locked': (action == 'lock')}
        )
        if action == 'unlock':
            from django.utils import timezone
            lock.last_unlocked_at = timezone.now()
            lock.save()
            
        messages.success(request, f"Estado {'bloqueado' if action == 'lock' else 'habilitado'} correctamente.")

        if next_url:
            return redirect(next_url)

    return redirect('manage_grade_locks')


@role_required('admin', 'director', 'teacher')
def academic_dashboard(request):
    context = {
        'total_courses': Course.objects.count(),
        'total_grade_records': GradeRecord.objects.count(),
        'total_periods': Period.objects.count(),
        'total_assignments': TeacherCourseAssignment.objects.count(),
    }
    return render(request, 'academic/academic_dashboard.html', context)


@role_required('admin', 'director')
def course_management(request):
    edit_course = None
    edit_id = request.GET.get('edit_course')
    if edit_id:
        edit_course = get_object_or_404(Course, id=edit_id)

    if request.method == 'POST':
        course_form = CourseForm(request.POST, instance=edit_course)
        if course_form.is_valid():
            course = course_form.save()
            if edit_course:
                messages.success(request, f"Curso '{course.name}' actualizado.")
            else:
                messages.success(request, f"Curso '{course.name}' creado correctamente.")
            return redirect('course_management')
    else:
        course_form = CourseForm(instance=edit_course)

    courses = Course.objects.order_by('name')
    context = {
        'course_form': course_form,
        'courses': courses,
        'edit_course': edit_course,
    }
    return render(request, 'academic/course_management.html', context)


@role_required('admin', 'director')
def delete_course(request, course_id):
    if request.method != 'POST':
        return redirect('course_management')

    course = get_object_or_404(Course, id=course_id)
    blockers = []

    if course.teachercourseassignment_set.exists():
        blockers.append('asignaciones de docentes')
    if course.graderecord_set.exists():
        blockers.append('registro de notas')
    if course.gradesubmissionlock_set.exists():
        blockers.append('bloqueos de notas')
    if course.poly_teachers_assigned.exists() or course.poly_teachers.exists():
        blockers.append('docentes polidocentes vinculados')

    if blockers:
        blocker_text = ', '.join(blockers)
        messages.error(
            request,
            f"No se puede eliminar el curso '{course.name}' porque tiene {blocker_text} asociados.",
        )
        return redirect('course_management')

    course_name = course.name
    course.delete()
    messages.success(request, f"Curso '{course_name}' eliminado correctamente.")
    return redirect('course_management')


@role_required('admin', 'director')
def course_grade_matrix(request):
    grades = list(Grade.objects.order_by('name'))
    courses = list(Course.objects.order_by('name'))

    if request.method == 'POST':
        posted = set(request.POST.keys())
        with transaction.atomic():
            for course in courses:
                selected_grade_ids = [
                    grade.id
                    for grade in grades
                    if f"cg_{course.id}_{grade.id}" in posted
                ]
                course.grades.set(selected_grade_ids)
            sync_teacher_course_assignments()

        messages.success(request, "Cursos por grado actualizados correctamente.")
        return redirect('course_grade_matrix')

    current = {
        course.id: set(course.grades.values_list('id', flat=True))
        for course in courses
    }
    for course in courses:
        course.selected_grade_ids = current.get(course.id, set())
    return render(request, 'academic/course_grade_matrix.html', {
        'grades': grades,
        'courses': courses,
    })


@role_required('admin', 'director')
def grade_management(request):
    edit_grade = None
    edit_id = request.GET.get('edit_grade')
    if edit_id:
        edit_grade = get_object_or_404(Grade, id=edit_id)

    if request.method == 'POST':
        form = GradeForm(request.POST, instance=edit_grade)
        if form.is_valid():
            grade = form.save()
            if edit_grade:
                messages.success(request, f"Grado '{grade.name}' actualizado.")
            else:
                messages.success(request, f"Grado '{grade.name}' creado correctamente.")
            return redirect('grade_management')
    else:
        form = GradeForm(instance=edit_grade)

    grades = Grade.objects.order_by('name')
    return render(request, 'academic/grade_management.html', {
        'form': form,
        'grades': grades,
        'edit_grade': edit_grade,
    })


@role_required('admin', 'director')
def section_management(request):
    edit_section = None
    previous_tutor_id = None
    edit_id = request.GET.get('edit_section')
    if edit_id:
        edit_section = get_object_or_404(Section, id=edit_id)
        previous_tutor_id = edit_section.tutor_teacher_id

    if request.method == 'POST':
        form = SectionForm(request.POST, instance=edit_section)
        if form.is_valid():
            section = form.save()
            tutor = section.tutor_teacher
            teachers_to_sync = set()
            stale_sections_by_teacher = {}
            if tutor:
                # Sync tutor assignment with teacher classroom profile.
                user_updates = {}
                if tutor.teaching_grade_id != section.grade_id:
                    user_updates['teaching_grade_id'] = section.grade_id
                if tutor.teaching_section_id != section.id:
                    user_updates['teaching_section_id'] = section.id
                if user_updates:
                    User.objects.filter(pk=tutor.pk).update(**user_updates)
                teachers_to_sync.add(tutor.pk)

            if previous_tutor_id and previous_tutor_id != section.tutor_teacher_id:
                # If previous tutor pointed to this section, clear stale profile section.
                User.objects.filter(
                    pk=previous_tutor_id,
                    teaching_section_id=section.id,
                ).update(teaching_section=None, teaching_grade=None)
                teachers_to_sync.add(previous_tutor_id)
                stale_sections_by_teacher[previous_tutor_id] = [section.id]

            if teachers_to_sync:
                sync_teacher_course_assignments(
                    teacher_ids=list(teachers_to_sync),
                    extra_section_ids_by_teacher=stale_sections_by_teacher,
                )

            if edit_section:
                messages.success(request, f"Sección '{section.name}' actualizada.")
            else:
                messages.success(request, f"Sección '{section.name}' creada para '{section.grade.name}'.")
            return redirect('section_management')
    else:
        form = SectionForm(instance=edit_section)

    sections = Section.objects.select_related('grade', 'tutor_teacher').order_by('grade__name', 'name')
    return render(request, 'academic/section_management.html', {
        'form': form,
        'sections': sections,
        'edit_section': edit_section,
    })


@role_required('admin', 'director', 'teacher')
def grade_list(request):
    student_order = resolve_student_order(request)
    grades = GradeRecord.objects.select_related(
        'enrollment__student',
        'course',
        'period'
    )

    if request.user.role == 'teacher' and not request.user.is_superuser:
        # Filter grades by courses assigned to the teacher
        assigned_course_ids = TeacherCourseAssignment.objects.filter(
            teacher=request.user
        ).values_list('course_id', flat=True)
        grades = grades.filter(course_id__in=assigned_course_ids)

    grades = order_queryset_by_student_name(
        grades,
        prefix='enrollment__student',
        student_order=student_order,
        extra_fields=['course__name', 'period__name', 'id'],
    )

    context = {
        'grades': grades,
    }
    context.update(student_order_context(request, student_order))
    return render(request, 'academic/grade_list.html', context)


@role_required('admin', 'director', 'teacher')
def grade_create(request):
    if request.method == 'POST':
        form = GradeRecordForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('grade_list')
    else:
        form = GradeRecordForm(user=request.user)

    return render(request, 'academic/grade_form.html', {
        'form': form
    })


@role_required('admin', 'director', 'teacher')
def report_card(request):
    # El reporte general siempre se consulta por grado y seccion.
    return redirect('course_report')


def _preferred_period(academic_year):
    if not academic_year:
        return None

    periods = list(
        Period.objects.filter(academic_year=academic_year).order_by('start_date', 'name')
    )
    if not periods:
        return None

    active = next((p for p in periods if p.is_active), None)
    if active:
        return active

    # Prefer "Bimestre 1" when present.
    for period in periods:
        normalized = (period.name or '').lower().replace(' ', '')
        if 'bimestre1' in normalized:
            return period

    # Otherwise, pick the earliest "bimestre".
    for period in periods:
        if 'bimestre' in (period.name or '').lower():
            return period

    return periods[0]


def _default_unit_name():
    return 'Unidad 1'


def _ensure_assignment_units(assignment, period=None):
    units_qs = assignment.units.select_related('period').order_by('period__start_date', 'order', 'id')
    if period:
        units_qs = units_qs.filter(period=period)
    units = list(units_qs)
    if not assignment.academic_year_id or not period:
        return units

    academic_periods = list(
        Period.objects.filter(academic_year=assignment.academic_year).order_by('start_date', 'name')
    )
    try:
        period_index = academic_periods.index(period)
    except ValueError:
        return units

    expected_orders = [period_index * 2 + 1, period_index * 2 + 2]
    existing_orders = {unit.order for unit in units}
    for order in expected_orders:
        if order not in existing_orders:
            units.append(
                Unit.objects.create(
                    assignment=assignment,
                    period=period,
                    name=f'Unidad {order}',
                    order=order,
                )
            )

    return sorted(units, key=lambda unit: (unit.order, unit.id))


def _is_section_tutor(user, section):
    return bool(
        user
        and section
        and getattr(user, 'role', None) == 'teacher'
        and not getattr(user, 'is_superuser', False)
        and section.tutor_teacher_id == user.id
    )


def _courses_for_enrollment(user, enrollment):
    if not enrollment:
        return Course.objects.none()

    is_teacher = user.role == 'teacher' and not user.is_superuser
    is_subject_teacher = is_teacher and not _is_section_tutor(user, enrollment.section)

    course_ids = set()
    if not is_subject_teacher:
        course_ids.update(
            enrollment.section.grade.courses.values_list('id', flat=True)
        )

    scope = TeacherCourseAssignment.objects.all()
    if enrollment.academic_year_id:
        scope = scope.filter(academic_year=enrollment.academic_year)

    scope = scope.filter(
        models.Q(section=enrollment.section)
        | models.Q(section__isnull=True, grade=enrollment.section.grade)
    )

    # Subject teachers only see their own courses; the tutor sees all courses of the section.
    if is_subject_teacher:
        scope = scope.filter(teacher=user)

    course_ids.update(scope.values_list('course_id', flat=True).distinct())
    if not course_ids:
        course_ids.update(
            GradeRecord.objects.filter(enrollment=enrollment)
            .values_list('course_id', flat=True)
            .distinct()
        )

    if not course_ids:
        return Course.objects.none()

    return Course.objects.filter(id__in=course_ids).order_by('name')


@role_required('admin', 'director', 'teacher', 'parent')
def student_report(request, enrollment_id):
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'section__grade', 'academic_year'),
        id=enrollment_id
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if not _is_section_tutor(request.user, enrollment.section):
            messages.error(request, "Solo la tutora del aula puede generar/ver la libreta del estudiante.")
            return redirect('teacher_dashboard')
    courses = list(_courses_for_enrollment(request.user, enrollment))
    period_id = _safe_int(request.GET.get('period'))
    preferred_period = _preferred_period(enrollment.academic_year)
    if period_id and enrollment.academic_year_id:
        preferred_period = (
            Period.objects.filter(academic_year=enrollment.academic_year, id=period_id).first()
            or preferred_period
        )
    _, breakdown = _build_competency_breakdown([enrollment], courses, enrollment.academic_year)

    course_cards = []
    for course in courses:
        course_data = breakdown.get((enrollment.id, course.id), {}) or {}
        period_finals = course_data.get('period_finals', {}) or {}

        if preferred_period:
            course_grade = period_finals.get(preferred_period.id, '-') or '-'
        else:
            course_grade = course_data.get('course_final', '-') or '-'

        competency_rows = []
        for competency_row in course_data.get('competencies', []) or []:
            if preferred_period:
                competency_grade = competency_row.get('period_grades', {}).get(preferred_period.id, '-') or '-'
            else:
                competency_grade = competency_row.get('final_grade', '-') or '-'
            competency_rows.append({
                'name': competency_row.get('name', '-') or '-',
                'grade': competency_grade,
            })

        if competency_rows or course_grade != '-':
            course_cards.append({
                'course': course,
                'course_grade': course_grade,
                'competencies': competency_rows,
            })

    return render(request, 'academic/student_report.html', {
        'enrollment': enrollment,
        'preferred_period': preferred_period,
        'courses': course_cards,
    })


@role_required('admin', 'director', 'teacher')
def course_report(request):
    grade_id = _safe_int(request.GET.get('grade'))
    section_id = _safe_int(request.GET.get('section'))
    student_order = resolve_student_order(request)
    report_data = _build_grade_section_report(
        request.user,
        grade_id=grade_id,
        section_id=section_id,
        student_order=student_order,
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if not report_data.get('selected_section'):
            messages.error(request, "Solo las tutoras pueden ver el reporte de notas por grado y seccion.")
            return redirect('teacher_dashboard')
    report_data.update(student_order_context(request, student_order))
    return render(request, 'academic/course_report.html', report_data)


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _active_year():
    return AcademicYear.objects.filter(is_active=True).order_by('-year').first()


def _courses_for_user(user):
    if user.role == 'teacher' and not user.is_superuser:
        assigned_course_ids = TeacherCourseAssignment.objects.filter(
            teacher=user
        ).values_list('course_id', flat=True)
        return Course.objects.filter(id__in=assigned_course_ids).order_by('name')
    return Course.objects.order_by('name')


def _is_bimestre_period_name(period_name):
    return bool(period_name and 'bimestre' in period_name.lower())


def _mode_from_period_grade_map(period_grade_map, period_name_by_id):
    all_grades = []
    bimestre_grades = []
    for period_id, grade in period_grade_map.items():
        if grade in {'AD', 'A', 'B', 'C'}:
            all_grades.append(grade)
            if _is_bimestre_period_name(period_name_by_id.get(period_id, '')):
                bimestre_grades.append(grade)
    return calculate_mode_grade(bimestre_grades or all_grades) or '-'


def _get_assignment_for_enrollment_and_course(enrollment, course, assignment_cache=None):
    if not enrollment or not course or not enrollment.academic_year_id:
        return None

    cache_key = (enrollment.academic_year_id, enrollment.section_id, course.id)
    if assignment_cache is not None and cache_key in assignment_cache:
        return assignment_cache[cache_key]

    base_qs = TeacherCourseAssignment.objects.select_related(
        'teacher',
        'course',
        'section__grade',
        'academic_year',
    ).prefetch_related('competencies__indicator_set').filter(
        course=course,
        academic_year=enrollment.academic_year,
    )

    assignment = None
    if enrollment.section_id:
        assignment = base_qs.filter(section=enrollment.section).first()
        if not assignment and enrollment.section.grade_id:
            assignment = base_qs.filter(
                section__isnull=True,
                grade=enrollment.section.grade,
            ).first()

    if not assignment:
        assignment = base_qs.filter(section__isnull=True).first()

    if assignment_cache is not None:
        assignment_cache[cache_key] = assignment

    return assignment


def _build_assignment_competency_breakdown(assignment, enrollments, periods, period_name_by_id, grade_record_map):
    competencies = list(assignment.competencies.all()) if assignment else []
    indicator_to_competency = {}
    competency_name_by_id = {}
    indicator_ids = []
    for competency in competencies:
        competency_name_by_id[competency.id] = competency.name
        for indicator in competency.indicator_set.all():
            indicator_to_competency[indicator.id] = competency.id
            indicator_ids.append(indicator.id)

    competency_period_raw = defaultdict(list)
    if indicator_ids:
        indicator_grades = IndicatorGrade.objects.filter(
            enrollment__in=enrollments,
            indicator_id__in=indicator_ids,
        )
        if periods:
            indicator_grades = indicator_grades.filter(period_id__in=period_name_by_id.keys())
        elif enrollments and enrollments[0].academic_year_id:
            indicator_grades = indicator_grades.filter(period__academic_year_id=enrollments[0].academic_year_id)
        for enrollment_id, indicator_id, period_id, grade in indicator_grades.values_list(
            'enrollment_id', 'indicator_id', 'period_id', 'grade'
        ):
            competency_id = indicator_to_competency.get(indicator_id)
            if competency_id:
                competency_period_raw[(enrollment_id, competency_id, period_id)].append(grade)

    competency_period_grade = {
        key: calculate_mode_grade(values)
        for key, values in competency_period_raw.items()
    }

    breakdown_by_enrollment = {}
    for enrollment in enrollments:
        competency_rows = []
        per_period_competency_grades = defaultdict(list)

        for competency in competencies:
            period_grades = {}
            for period in periods:
                grade = competency_period_grade.get((enrollment.id, competency.id, period.id))
                if grade:
                    period_grades[period.id] = grade
                    per_period_competency_grades[period.id].append(grade)

            competency_rows.append({
                'name': competency_name_by_id.get(competency.id, '-'),
                'period_grades': period_grades,
                'final_grade': _mode_from_period_grade_map(period_grades, period_name_by_id),
            })

        period_finals = {}
        for period in periods:
            comp_grades = per_period_competency_grades.get(period.id, [])
            period_finals[period.id] = (
                calculate_mode_grade(comp_grades)
                or grade_record_map.get((enrollment.id, assignment.course_id, period.id))
                or '-'
            )

        course_final = _mode_from_period_grade_map(
            {period_id: grade for period_id, grade in period_finals.items() if grade != '-'},
            period_name_by_id,
        )

        breakdown_by_enrollment[enrollment.id] = {
            'competencies': competency_rows,
            'period_finals': period_finals,
            'course_final': course_final,
        }

    return breakdown_by_enrollment


def _build_competency_breakdown(enrollments, courses, academic_year=None):
    enrollments = list(enrollments)
    courses = list(courses)
    if not enrollments or not courses:
        return [], {}

    periods_qs = Period.objects.all()
    if academic_year:
        periods_qs = periods_qs.filter(academic_year=academic_year)
    periods = list(periods_qs.order_by('start_date', 'name'))
    period_name_by_id = {period.id: period.name for period in periods}

    grade_record_map = {}
    grade_records = GradeRecord.objects.filter(
        enrollment__in=enrollments,
        course__in=courses,
    )
    if periods:
        grade_records = grade_records.filter(period_id__in=period_name_by_id.keys())
    elif academic_year:
        grade_records = grade_records.filter(period__academic_year=academic_year)
    for enrollment_id, course_id, period_id, grade in grade_records.values_list(
        'enrollment_id', 'course_id', 'period_id', 'grade'
    ):
        grade_record_map[(enrollment_id, course_id, period_id)] = grade

    assignment_cache = {}
    assignment_breakdown_cache = {}
    breakdown = {}
    for enrollment in enrollments:
        for course in courses:
            assignment = _get_assignment_for_enrollment_and_course(enrollment, course, assignment_cache)
            if assignment:
                assignment_breakdown = assignment_breakdown_cache.get(assignment.id)
                if assignment_breakdown is None:
                    assignment_breakdown = _build_assignment_competency_breakdown(
                        assignment,
                        enrollments,
                        periods,
                        period_name_by_id,
                        grade_record_map,
                    )
                    assignment_breakdown_cache[assignment.id] = assignment_breakdown
                breakdown[(enrollment.id, course.id)] = assignment_breakdown.get(enrollment.id, {
                    'competencies': [],
                    'period_finals': {
                        period.id: grade_record_map.get((enrollment.id, course.id, period.id), '-')
                        for period in periods
                    },
                    'course_final': _mode_from_period_grade_map(
                        {
                            period.id: grade_record_map.get((enrollment.id, course.id, period.id))
                            for period in periods
                            if grade_record_map.get((enrollment.id, course.id, period.id))
                        },
                        period_name_by_id,
                    ),
                })
                continue

            period_finals = {
                period.id: grade_record_map.get((enrollment.id, course.id, period.id), '-')
                for period in periods
            }
            breakdown[(enrollment.id, course.id)] = {
                'competencies': [],
                'period_finals': period_finals,
                'course_final': _mode_from_period_grade_map(
                    {
                        period_id: grade
                        for period_id, grade in period_finals.items()
                        if grade != '-'
                    },
                    period_name_by_id,
                ),
            }

    return periods, breakdown


def _build_grade_section_report(user, grade_id=None, section_id=None, student_order='az'):
    is_teacher = user.role == 'teacher' and not user.is_superuser
    active_year = _active_year()

    grades = Grade.objects.order_by('name')
    sections = Section.objects.select_related('grade').order_by('grade__name', 'name')

    assignment_scope = TeacherCourseAssignment.objects.select_related('section__grade', 'grade')
    if active_year:
        assignment_scope = assignment_scope.filter(academic_year=active_year)

    if is_teacher:
        # Only the classroom tutor can view the consolidated report for that section.
        sections = sections.filter(tutor_teacher=user)
        grade_option_ids = list(sections.values_list('grade_id', flat=True).distinct())
        grades = grades.filter(id__in=grade_option_ids) if grade_option_ids else grades.none()

    selected_grade = grades.filter(id=grade_id).first() if grade_id else None
    if not selected_grade:
        selected_grade = grades.first()

    sections_for_grade = sections.filter(grade=selected_grade) if selected_grade else sections.none()
    selected_section = sections_for_grade.filter(id=section_id).first() if section_id else None
    if selected_grade and not selected_section:
        selected_section = sections_for_grade.first()

    rows = []
    courses = Course.objects.none()

    if selected_grade and selected_section:
        enrollments = Enrollment.objects.select_related(
            'student',
            'section__grade',
        ).filter(
            status='active',
            section__grade=selected_grade,
        )
        if active_year:
            enrollments = enrollments.filter(academic_year=active_year)
        enrollments = enrollments.filter(section=selected_section)

        if is_teacher:
            enrollments = enrollments.filter(section=selected_section)

        enrollments = list(order_queryset_by_student_name(
            enrollments,
            prefix='student',
            student_order=student_order,
            extra_fields=['section__grade__name', 'section__name', 'id'],
        )
        )

        course_ids = set(selected_grade.courses.values_list('id', flat=True))

        section_course_scope = assignment_scope.filter(
            models.Q(section=selected_section)
            | models.Q(section__isnull=True, grade=selected_grade)
        )

        course_ids.update(section_course_scope.values_list('course_id', flat=True).distinct())
        if not course_ids and enrollments:
            course_ids.update(
                GradeRecord.objects.filter(enrollment__in=enrollments).values_list('course_id', flat=True).distinct()
            )
        courses = Course.objects.filter(id__in=course_ids).order_by('name')
        course_list = list(courses)
        _, breakdown = _build_competency_breakdown(enrollments, course_list, active_year)

        for enrollment in enrollments:
            rows.append({
                'enrollment_id': enrollment.id,
                'student': enrollment.student,
                'grade_name': enrollment.section.grade.name,
                'section_name': enrollment.section.name,
                'grades': [
                    breakdown.get((enrollment.id, course.id), {}).get('course_final', '-')
                    for course in course_list
                ],
            })

    return {
        'grades': grades,
        'sections': sections_for_grade,
        'selected_grade': selected_grade,
        'selected_section': selected_section,
        'rows': rows,
        'courses': courses,
        'active_year': active_year,
    }


@role_required('admin', 'director', 'teacher')
def course_report_export_excel(request):
    grade_id = _safe_int(request.GET.get('grade'))
    section_id = _safe_int(request.GET.get('section'))
    student_order = resolve_student_order(request)
    report_data = _build_grade_section_report(
        request.user,
        grade_id=grade_id,
        section_id=section_id,
        student_order=student_order,
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if not report_data.get('selected_section'):
            messages.error(request, "Solo las tutoras pueden exportar el reporte por grado y seccion.")
            return redirect('teacher_dashboard')

    from openpyxl import Workbook
    from openpyxl.styles import Font

    course_list = list(report_data['courses'])
    enrollment_ids = [row['enrollment_id'] for row in report_data['rows']]
    enrollment_map = {
        enrollment.id: enrollment
        for enrollment in Enrollment.objects.select_related('student', 'section__grade').filter(id__in=enrollment_ids)
    }
    enrollments = [enrollment_map[eid] for eid in enrollment_ids if eid in enrollment_map]
    periods, breakdown = _build_competency_breakdown(enrollments, course_list, report_data['active_year'])

    def _adjust_width(sheet):
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    headers = ['Alumno', 'Grado', 'Seccion'] + [course.name for course in course_list]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in report_data['rows']:
        ws.append([
            str(row['student']),
            row['grade_name'],
            row['section_name'],
            *row['grades'],
        ])

    _adjust_width(ws)

    details_ws = wb.create_sheet(title="Competencias")
    details_headers = ['Alumno', 'Grado', 'Seccion', 'Curso', 'Competencia'] + [period.name for period in periods] + ['Final competencia']
    details_ws.append(details_headers)
    for cell in details_ws[1]:
        cell.font = Font(bold=True)

    for enrollment in enrollments:
        for course in course_list:
            course_data = breakdown.get((enrollment.id, course.id), {})
            competency_rows = course_data.get('competencies', [])
            if not competency_rows:
                details_ws.append([
                    str(enrollment.student),
                    enrollment.section.grade.name,
                    enrollment.section.name,
                    course.name,
                    'Sin competencias',
                    *(('-' for _ in periods)),
                    '-',
                ])
                continue

            for competency_row in competency_rows:
                details_ws.append([
                    str(enrollment.student),
                    enrollment.section.grade.name,
                    enrollment.section.name,
                    course.name,
                    competency_row['name'],
                    *[competency_row['period_grades'].get(period.id, '-') for period in periods],
                    competency_row.get('final_grade', '-') or '-',
                ])
    _adjust_width(details_ws)

    period_ws = wb.create_sheet(title="Finales por periodo")
    period_headers = ['Alumno', 'Grado', 'Seccion', 'Curso'] + [period.name for period in periods] + ['Final curso']
    period_ws.append(period_headers)
    for cell in period_ws[1]:
        cell.font = Font(bold=True)

    for enrollment in enrollments:
        for course in course_list:
            course_data = breakdown.get((enrollment.id, course.id), {})
            period_finals = course_data.get('period_finals', {})
            period_ws.append([
                str(enrollment.student),
                enrollment.section.grade.name,
                enrollment.section.name,
                course.name,
                *[period_finals.get(period.id, '-') for period in periods],
                course_data.get('course_final', '-') or '-',
            ])
    _adjust_width(period_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    grade_slug = report_data['selected_grade'].name.replace(' ', '_') if report_data['selected_grade'] else 'sin_grado'
    section_slug = report_data['selected_section'].name.replace(' ', '_') if report_data['selected_section'] else 'todas_secciones'
    filename = f"reporte_notas_{grade_slug}_{section_slug}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('admin', 'director', 'teacher')
def period_report(request):
    student_order = resolve_student_order(request)
    periods = Period.objects.select_related('academic_year').order_by('-academic_year__year', 'start_date', 'name')
    selected_period = None
    grades = []
    period_id = request.GET.get('period')
    is_teacher = request.user.role == 'teacher' and not request.user.is_superuser

    if period_id:
        selected_period = get_object_or_404(Period, id=period_id)
        grades_query = GradeRecord.objects.select_related(
            'enrollment__student',
            'course',
            'period'
        ).filter(period=selected_period)
        
        if is_teacher:
            assigned_course_ids = TeacherCourseAssignment.objects.filter(teacher=request.user).values_list('course_id', flat=True)
            grades_query = grades_query.filter(course_id__in=assigned_course_ids)

        grades = order_queryset_by_student_name(
            grades_query,
            prefix='enrollment__student',
            student_order=student_order,
            extra_fields=['course__name', 'id'],
        )

    context = {
        'periods': periods,
        'selected_period': selected_period,
        'grades': grades,
    }
    context.update(student_order_context(request, student_order))
    return render(request, 'academic/period_report.html', context)


@role_required('admin', 'director', 'teacher', 'parent')
def student_report_pdf(request, enrollment_id):
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'section__grade', 'academic_year__school'),
        id=enrollment_id
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if not _is_section_tutor(request.user, enrollment.section):
            messages.error(request, "Solo la tutora del aula puede exportar la libreta del estudiante.")
            return redirect('teacher_dashboard')
    courses = list(_courses_for_enrollment(request.user, enrollment))
    period_id = _safe_int(request.GET.get('period'))
    preferred_period = _preferred_period(enrollment.academic_year)
    if period_id and enrollment.academic_year_id:
        preferred_period = (
            Period.objects.filter(academic_year=enrollment.academic_year, id=period_id).first()
            or preferred_period
        )
    _, breakdown = _build_competency_breakdown([enrollment], courses, enrollment.academic_year)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            "No se pudo generar PDF porque falta la libreria 'reportlab'.",
            content_type='text/plain'
        )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=14,
        alignment=1,  # center
    )
    style_cell = ParagraphStyle(
        name='Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.6,
        leading=9.4,
    )
    style_cell_bold = ParagraphStyle(
        name='CellBold',
        parent=style_cell,
        fontName='Helvetica-Bold',
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
        title="Boleta Academica",
    )

    story = []
    school = getattr(enrollment.academic_year, 'school', None) if enrollment.academic_year_id else None
    year_label = str(enrollment.academic_year.year) if enrollment.academic_year_id else ''

    def _maybe_logo(max_w=52, max_h=52):
        if not school or not getattr(school, 'logo', None):
            return None
        try:
            logo_path = school.logo.path
        except Exception:
            return None
        try:
            img_reader = ImageReader(logo_path)
            iw, ih = img_reader.getSize()
        except Exception:
            return None
        if not iw or not ih:
            return None
        scale = min(max_w / float(iw), max_h / float(ih))
        return Image(logo_path, width=iw * scale, height=ih * scale)

    left_logo = _maybe_logo()
    right_logo = _maybe_logo()

    header = Table(
        [[
            left_logo or '',
            Paragraph(f"INFORME DE PROGRESO DE APRENDIZAJE DEL ESTUDIANTE - {year_label}", style_title),
            right_logo or '',
        ]],
        colWidths=[60, doc.width - 120, 60],
    )
    header.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#D9F0FF')),
        ('BOX', (1, 0), (1, 0), 0.8, colors.HexColor('#6AAED6')),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(header)
    story.append(Spacer(1, 8))

    student = enrollment.student
    student_full_name = f"{student.last_name}, {student.first_name}".upper() if student else "-"
    grade_name = str(enrollment.section.grade) if enrollment.section_id else "-"
    section_name = str(enrollment.section.name) if enrollment.section_id else "-"
    school_name = str(school.name) if school else "-"

    info = [
        [Paragraph("NIVEL", style_cell_bold), Paragraph("Primaria", style_cell),
         Paragraph("I.E.", style_cell_bold), Paragraph(school_name, style_cell)],
        [Paragraph("GRADO", style_cell_bold), Paragraph(str(grade_name), style_cell),
         Paragraph("SECCION", style_cell_bold), Paragraph(str(section_name), style_cell)],
        [Paragraph("APELLIDOS Y NOMBRES", style_cell_bold), Paragraph(student_full_name, style_cell_bold), "", ""],
    ]
    info_table = Table(info, colWidths=[90, 150, 90, doc.width - 330])
    info_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#D9F0FF')),
        ('SPAN', (1, 2), (-1, 2)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 10))

    bimestre_periods = []
    if enrollment.academic_year_id:
        bimestre_periods = [
            p for p in Period.objects.filter(academic_year=enrollment.academic_year).order_by('start_date', 'name')
            if 'bimestre' in (p.name or '').lower()
        ][:4]

    # Map columns I..IV to bimestre periods (if missing, it will show "-").
    period_cols = (bimestre_periods + [None, None, None, None])[:4]

    big_table_data = [
        [
            Paragraph("AREAS", style_cell_bold),
            Paragraph("COMPETENCIAS", style_cell_bold),
            Paragraph("CALIFICATIVO POR BIMESTRE", style_cell_bold),
            "",
            "",
            "",
            Paragraph("PF", style_cell_bold),
        ],
        [
            "",
            "",
            Paragraph("I", style_cell_bold),
            Paragraph("II", style_cell_bold),
            Paragraph("III", style_cell_bold),
            Paragraph("IV", style_cell_bold),
            "",
        ],
    ]

    area_spans = []
    current_row = 2  # after headers

    for course in courses:
        course_data = breakdown.get((enrollment.id, course.id), {}) or {}
        competencies = course_data.get('competencies', []) or []
        if not competencies:
            continue

        start_row = current_row
        for competency_row in competencies:
            period_grades = competency_row.get('period_grades', {}) or {}
            row = [
                Paragraph(str(course.name).upper(), style_cell_bold),
                Paragraph(str(competency_row.get('name', '-') or '-'), style_cell),
            ]
            for period in period_cols:
                if not period:
                    row.append(Paragraph("-", style_cell))
                    continue
                row.append(Paragraph(str(period_grades.get(period.id, '-') or '-'), style_cell_bold))
            row.append(Paragraph(str(competency_row.get('final_grade', '-') or '-'), style_cell_bold))
            big_table_data.append(row)
            current_row += 1

        end_row = current_row - 1
        if end_row > start_row:
            area_spans.append((start_row, end_row))

    col_widths = [
        78,  # areas
        doc.width - (78 + 4 * 36 + 36),  # competencies
        36, 36, 36, 36,  # I-IV
        36,  # PF
    ]
    t = Table(big_table_data, colWidths=col_widths, repeatRows=2)
    table_style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#D9F0FF')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (2, 2), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('SPAN', (0, 0), (0, 1)),
        ('SPAN', (1, 0), (1, 1)),
        ('SPAN', (2, 0), (5, 0)),
        ('SPAN', (6, 0), (6, 1)),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    for start_row, end_row in area_spans:
        table_style_cmds.append(('SPAN', (0, start_row), (0, end_row)))
        table_style_cmds.append(('VALIGN', (0, start_row), (0, end_row), 'MIDDLE'))
        table_style_cmds.append(('ALIGN', (0, start_row), (0, end_row), 'CENTER'))

    t.setStyle(TableStyle(table_style_cmds))
    story.append(t)

    doc.build(story)
    pdf_bytes = buffer.getvalue()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="boleta_{enrollment.id}.pdf"'
    return response


def _calculate_course_grade_from_indicators(enrollment, assignment, period):
    competencies = list(assignment.competencies.prefetch_related('indicator_set').all()) if assignment else []
    competency_grades = []
    for competency in competencies:
        indicator_ids = list(competency.indicator_set.values_list('id', flat=True))
        if not indicator_ids:
            continue
        grades = IndicatorGrade.objects.filter(
            enrollment=enrollment,
            period=period,
            indicator_id__in=indicator_ids
        ).values_list('grade', flat=True)
        comp_grade = calculate_mode_grade(list(grades))
        if comp_grade:
            competency_grades.append(comp_grade)

    return calculate_mode_grade(competency_grades)


def _replicate_indicators_to_unit(competency, source_unit, target_unit):
    if not competency or not source_unit or not target_unit:
        return 0

    created_count = 0
    source_indicators = Indicator.objects.filter(
        competency=competency,
        unit=source_unit,
    ).order_by('order', 'id')

    for source_indicator in source_indicators:
        exists = Indicator.objects.filter(
            competency=competency,
            unit=target_unit,
            name=source_indicator.name,
        ).exists()
        if exists:
            continue

        Indicator.objects.create(
            competency=competency,
            unit=target_unit,
            name=source_indicator.name,
            order=source_indicator.order,
        )
        created_count += 1

    return created_count


@role_required('admin', 'director', 'teacher')
def teacher_competency_gradebook(request):
    student_order = resolve_student_order(request)
    assignments = TeacherCourseAssignment.objects.select_related(
        'teacher',
        'course',
        'section__grade',
        'academic_year'
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        assignments = assignments.filter(teacher=request.user)

    assignments = assignments.order_by(
        'course__name',
        'section__grade__name',
        'section__name',
        'id',
    )
    assignment_options = [
        {
            'assignment': assignment,
            'label': _assignment_select_label(assignment),
        }
        for assignment in assignments
    ]

    active_year = AcademicYear.objects.filter(is_active=True).first()
    periods = (
        Period.objects.filter(academic_year=active_year).order_by('start_date', 'name')
        if active_year else Period.objects.none()
    )

    assignment_id = request.POST.get('assignment') if request.method == 'POST' else request.GET.get('assignment')
    selected_assignment = assignments.filter(id=assignment_id).first() if assignment_id else None

    selected_period = None
    period_id = request.POST.get('period') if request.method == 'POST' else request.GET.get('period')
    if selected_assignment:
        if active_year != selected_assignment.academic_year:
            periods = Period.objects.filter(
                academic_year=selected_assignment.academic_year
            ).order_by('start_date', 'name')
        selected_period = periods.filter(id=period_id).first() if period_id else _preferred_period(selected_assignment.academic_year)

    units = []
    selected_unit = None
    unit_id = request.POST.get('unit') if request.method == 'POST' else request.GET.get('unit')
    if selected_assignment:
        units = _ensure_assignment_units(selected_assignment, selected_period)
        selected_unit = next((unit for unit in units if str(unit.id) == str(unit_id)), None) if unit_id else units[0]
        if not selected_unit and units:
            selected_unit = units[0]

    competencies = []
    competency_blocks = []
    rows = []
    competency_assignment = selected_assignment
    competency_unit = None
    is_locked = False
    can_edit = True
    can_unlock_lock = False
    lock_next_url = ''

    if selected_assignment and selected_period and selected_unit:
        competency_assignment = _competency_source_assignment(selected_assignment)
        if competency_assignment and competency_assignment.id != selected_assignment.id:
            source_units = _ensure_assignment_units(competency_assignment, selected_period)
            competency_unit = next(
                (
                    unit for unit in source_units
                    if unit.order == selected_unit.order
                    or unit.name == selected_unit.name
                ),
                None,
            )
        else:
            competency_unit = selected_unit
        competency_unit = competency_unit or selected_unit

        competencies = list(competency_assignment.competencies.all())
        for competency in competencies:
            indicators = list(
                competency.indicator_set.filter(unit=competency_unit).order_by('order', 'id')
            )
            competency_blocks.append({'competency': competency, 'indicators': indicators})

        enrollment_filters = {
            'academic_year': selected_assignment.academic_year,
            'status': 'active'
        }
        if selected_assignment.section:
            enrollment_filters['section'] = selected_assignment.section
        elif selected_assignment.grade:
            enrollment_filters['section__grade'] = selected_assignment.grade

        enrollments = order_queryset_by_student_name(
            Enrollment.objects.select_related('student').filter(**enrollment_filters),
            prefix='student',
            student_order=student_order,
        )

        all_indicator_ids = [indicator.id for block in competency_blocks for indicator in block['indicators']]
        score_values = {}
        if all_indicator_ids:
            score_values = {
                f"{score.enrollment_id}_{score.indicator_id}": score.grade
                for score in IndicatorGrade.objects.filter(
                    enrollment__in=enrollments,
                    period=selected_period,
                    indicator_id__in=all_indicator_ids
                )
            }

        lock = GradeSubmissionLock.objects.filter(
            teacher=selected_assignment.teacher,
            course=selected_assignment.course,
            section=selected_assignment.section,
            period=selected_period
        ).first()
        is_locked = lock.is_locked if lock else False
        can_edit = not is_locked
        can_unlock_lock = (
            is_locked and (
                request.user.role in {'admin', 'director'} or request.user.is_superuser
            )
        )
        lock_next_url = (
            f"{request.path}?assignment={selected_assignment.id}"
            f"&period={selected_period.id}&unit={selected_unit.id}&student_order={student_order}"
        )

        if request.method == 'POST':
            if is_locked:
                if can_unlock_lock:
                    messages.error(request, 'El registro esta bloqueado. Primero habilita la edicion.')
                else:
                    messages.error(request, 'Las notas para esta unidad estan bloqueadas. Contacte al director o admin.')
                return redirect(lock_next_url)

            valid_grades = {'AD', 'A', 'B', 'C'}
            with transaction.atomic():
                existing_records = {
                    (record.enrollment_id, record.indicator_id): record
                    for record in IndicatorGrade.objects.filter(
                        enrollment__in=enrollments,
                        period=selected_period,
                        indicator_id__in=all_indicator_ids
                    )
                }

                for enrollment in enrollments:
                    for block in competency_blocks:
                        for indicator in block['indicators']:
                            field_name = f"score_{enrollment.id}_{indicator.id}"
                            value = request.POST.get(field_name, '').strip().upper()
                            key = (enrollment.id, indicator.id)
                            current = existing_records.get(key)
                            if value in valid_grades:
                                if current:
                                    if current.grade != value:
                                        current.grade = value
                                        current.save(update_fields=['grade'])
                                else:
                                    IndicatorGrade.objects.create(
                                        enrollment=enrollment,
                                        indicator=indicator,
                                        period=selected_period,
                                        grade=value,
                                    )
                            elif current:
                                current.delete()

                for enrollment in enrollments:
                    final_grade = _calculate_course_grade_from_indicators(
                        enrollment, competency_assignment, selected_period
                    )
                    if final_grade:
                        GradeRecord.objects.update_or_create(
                            enrollment=enrollment,
                            course=selected_assignment.course,
                            period=selected_period,
                            defaults={'grade': final_grade}
                        )
                    else:
                        GradeRecord.objects.filter(
                            enrollment=enrollment,
                            course=selected_assignment.course,
                            period=selected_period
                        ).delete()

                if request.POST.get('finalize'):
                    lock_obj, _ = GradeSubmissionLock.objects.get_or_create(
                        teacher=selected_assignment.teacher,
                        course=selected_assignment.course,
                        section=selected_assignment.section,
                        period=selected_period
                    )
                    lock_obj.is_locked = True
                    lock_obj.save()
                    messages.success(request, 'Notas finalizadas y bloqueadas.')
                else:
                    messages.success(request, 'Borrador de notas guardado.')

            return redirect(lock_next_url)

        for enrollment in enrollments:
            blocks = []
            competency_grades = []
            for block in competency_blocks:
                cells = []
                indicator_grades = []
                for indicator in block['indicators']:
                    score_key = f"{enrollment.id}_{indicator.id}"
                    value = score_values.get(score_key, '')
                    if value:
                        indicator_grades.append(value)
                    cells.append({
                        'indicator': indicator,
                        'field_name': f"score_{score_key}",
                        'value': value,
                    })
                competency_grade = calculate_mode_grade(indicator_grades) or '-'
                if competency_grade != '-':
                    competency_grades.append(competency_grade)
                blocks.append({
                    'competency': block['competency'],
                    'cells': cells,
                    'competency_grade': competency_grade,
                })

            rows.append({
                'enrollment': enrollment,
                'blocks': blocks,
                'course_grade': calculate_mode_grade(competency_grades) or '-',
            })

        has_competencies = any(block['indicators'] for block in competency_blocks)

    context = {
        'assignment_options': assignment_options,
        'assignments': assignments,
        'selected_assignment': selected_assignment,
        'competency_assignment': competency_assignment,
        'periods': periods,
        'selected_period': selected_period,
        'units': units,
        'selected_unit': selected_unit,
        'competency_blocks': competency_blocks,
        'rows': rows,
        'grade_options': GradeRecord.GRADE_SCALE,
        'has_competencies': has_competencies if selected_assignment and selected_period and selected_unit else False,
        'is_locked': is_locked,
        'can_edit': can_edit,
        'can_unlock_lock': can_unlock_lock,
        'lock_next_url': lock_next_url,
    }
    context.update(student_order_context(request, student_order))
    return render(request, 'academic/teacher_gradebook.html', context)


@role_required('admin', 'director', 'teacher')
def teacher_competency_gradebook_export_excel(request):
    student_order = resolve_student_order(request)
    assignments = TeacherCourseAssignment.objects.select_related(
        'teacher',
        'course',
        'section__grade',
        'academic_year'
    )
    if request.user.role == 'teacher' and not request.user.is_superuser:
        assignments = assignments.filter(teacher=request.user)

    assignment_id = request.POST.get('assignment') if request.method == 'POST' else request.GET.get('assignment')
    selected_assignment = assignments.filter(id=assignment_id).first() if assignment_id else None
    if not selected_assignment:
        messages.error(request, 'Selecciona un curso antes de exportar.')
        return redirect('teacher_competency_gradebook')

    active_year = AcademicYear.objects.filter(is_active=True).first()
    periods = (
        Period.objects.filter(academic_year=active_year).order_by('start_date', 'name')
        if active_year else Period.objects.none()
    )

    period_id = request.POST.get('period') if request.method == 'POST' else request.GET.get('period')
    if active_year != selected_assignment.academic_year:
        periods = Period.objects.filter(
            academic_year=selected_assignment.academic_year
        ).order_by('start_date', 'name')
    selected_period = periods.filter(id=period_id).first() if period_id else _preferred_period(selected_assignment.academic_year)
    if not selected_period:
        messages.error(request, 'Selecciona un periodo valido antes de exportar.')
        return redirect('teacher_competency_gradebook')

    units = _ensure_assignment_units(selected_assignment, selected_period)
    unit_id = request.POST.get('unit') if request.method == 'POST' else request.GET.get('unit')
    selected_unit = next((unit for unit in units if str(unit.id) == str(unit_id)), None) if unit_id else units[0]
    if not selected_unit and units:
        selected_unit = units[0]
    if not selected_unit:
        messages.error(request, 'Selecciona una unidad valida antes de exportar.')
        return redirect('teacher_competency_gradebook')

    competency_assignment = _competency_source_assignment(selected_assignment)
    if competency_assignment and competency_assignment.id != selected_assignment.id:
        source_units = _ensure_assignment_units(competency_assignment, selected_period)
        competency_unit = next(
            (
                unit for unit in source_units
                if unit.order == selected_unit.order
                or unit.name == selected_unit.name
            ),
            None,
        )
    else:
        competency_unit = selected_unit
    competency_unit = competency_unit or selected_unit

    competency_blocks = []
    competencies = list(competency_assignment.competencies.all())
    for competency in competencies:
        indicators = list(
            competency.indicator_set.filter(unit=competency_unit).order_by('order', 'id')
        )
        competency_blocks.append({'competency': competency, 'indicators': indicators})

    enrollment_filters = {
        'academic_year': selected_assignment.academic_year,
        'status': 'active'
    }
    if selected_assignment.section:
        enrollment_filters['section'] = selected_assignment.section
    elif selected_assignment.grade:
        enrollment_filters['section__grade'] = selected_assignment.grade

    enrollments = order_queryset_by_student_name(
        Enrollment.objects.select_related('student', 'section__grade').filter(**enrollment_filters),
        prefix='student',
        student_order=student_order,
    )

    all_indicator_ids = [indicator.id for block in competency_blocks for indicator in block['indicators']]
    score_values = {}
    posted_score_values = {
        key.removeprefix('score_'): (value or '').strip().upper()
        for key, value in request.POST.items()
        if key.startswith('score_') and (value or '').strip().upper() in {'AD', 'A', 'B', 'C'}
    }
    if posted_score_values:
        score_values = posted_score_values
    elif all_indicator_ids:
        score_values = {
            f"{score.enrollment_id}_{score.indicator_id}": score.grade
            for score in IndicatorGrade.objects.filter(
                enrollment__in=enrollments,
                period=selected_period,
                indicator_id__in=all_indicator_ids
            )
        }

    rows = []
    for enrollment in enrollments:
        blocks = []
        competency_grades = []
        for block in competency_blocks:
            cells = []
            indicator_grades = []
            for indicator in block['indicators']:
                score_key = f"{enrollment.id}_{indicator.id}"
                value = score_values.get(score_key, '')
                if value:
                    indicator_grades.append(value)
                cells.append({
                    'indicator': indicator,
                    'field_name': f"score_{score_key}",
                    'value': value,
                })
            competency_grade = calculate_mode_grade(indicator_grades) or '-'
            if competency_grade != '-':
                competency_grades.append(competency_grade)
            blocks.append({
                'competency': block['competency'],
                'cells': cells,
                'competency_grade': competency_grade,
            })

        rows.append({
            'enrollment': enrollment,
            'blocks': blocks,
            'course_grade': calculate_mode_grade(competency_grades) or '-',
        })

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from django.utils.text import slugify

    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas'
    ws.freeze_panes = 'C7'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    last_column = 2 + sum(len(block['indicators']) + 1 for block in competency_blocks) + 1
    last_column_letter = ws.cell(row=1, column=last_column).column_letter
    header_font = Font(name='Arial', bold=True, size=11, color='000000')
    title_font = Font(name='Arial', bold=True, size=12, color='000000')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    vertical = Alignment(horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fills = [
        PatternFill('solid', fgColor='FFF2CC'),
        PatternFill('solid', fgColor='E2F0D9'),
        PatternFill('solid', fgColor='D9E2F3'),
        PatternFill('solid', fgColor='FCE4D6'),
    ]

    school_name = (
        getattr(getattr(selected_assignment.academic_year, 'school', None), 'name', None)
        or 'SISTEMA ESCOLAR'
    )
    classroom_name = (
        f"{selected_assignment.section.grade.name} {selected_assignment.section.name}"
        if selected_assignment.section else (selected_assignment.grade.name if selected_assignment.grade else 'Sin seccion')
    )
    area_name = f"AREA DE {selected_assignment.course.name.upper()}"
    subtitle = f"{selected_period.name} - {selected_unit.name}"

    ws.merge_cells(f"A1:{last_column_letter}1")
    ws['A1'] = school_name
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells(f"A2:{last_column_letter}2")
    ws['A2'] = 'REGISTRO AUXILIAR'
    ws['A2'].font = title_font
    ws['A2'].alignment = center

    ws.merge_cells(f"A3:{last_column_letter}3")
    ws['A3'] = subtitle
    ws['A3'].font = header_font
    ws['A3'].alignment = center

    ws.merge_cells(f"A4:{last_column_letter}4")
    ws['A4'] = f"{area_name} - {classroom_name}"
    ws['A4'].font = header_font
    ws['A4'].alignment = center

    ws.merge_cells('A5:A6')
    ws['A5'] = 'N'
    ws['A5'].font = header_font
    ws['A5'].alignment = center

    ws.merge_cells('B5:B6')
    ws['B5'] = 'NOMBRES Y APELLIDOS'
    ws['B5'].font = header_font
    ws['B5'].alignment = center

    current_column = 3
    for block_index, block in enumerate(competency_blocks):
        block_start = current_column
        block_end = current_column + len(block['indicators'])
        fill = fills[block_index % len(fills)]
        ws.merge_cells(start_row=5, start_column=block_start, end_row=5, end_column=block_end)
        top_cell = ws.cell(row=5, column=block_start)
        top_cell.value = block['competency'].name
        top_cell.font = header_font
        top_cell.alignment = center
        for column in range(block_start, block_end + 1):
            ws.cell(row=5, column=column).fill = fill
            ws.cell(row=6, column=column).fill = fill

        for offset, indicator in enumerate(block['indicators']):
            indicator_cell = ws.cell(row=6, column=current_column + offset)
            indicator_cell.value = indicator.name
            indicator_cell.font = header_font
            indicator_cell.alignment = vertical
            indicator_cell.border = thin_border

        promedio_cell = ws.cell(row=6, column=block_end)
        promedio_cell.value = 'PROMEDIO'
        promedio_cell.font = header_font
        promedio_cell.alignment = vertical
        promedio_cell.border = thin_border
        current_column = block_end + 1

    course_cell = ws.cell(row=5, column=current_column)
    course_cell.value = 'NOTA CURSO'
    course_cell.font = header_font
    course_cell.alignment = center
    ws.merge_cells(start_row=5, start_column=current_column, end_row=6, end_column=current_column)
    ws.cell(row=5, column=current_column).fill = PatternFill('solid', fgColor='D9EAD3')
    ws.cell(row=6, column=current_column).fill = PatternFill('solid', fgColor='D9EAD3')

    for row_index, row in enumerate(rows, start=7):
        ws.cell(row=row_index, column=1, value=row_index - 6)
        ws.cell(row=row_index, column=2, value=str(row['enrollment'].student))
        ws.cell(row=row_index, column=1).alignment = center
        ws.cell(row=row_index, column=2).alignment = Alignment(vertical='center')
        ws.cell(row=row_index, column=1).border = thin_border
        ws.cell(row=row_index, column=2).border = thin_border

        column_index = 3
        for block in row['blocks']:
            for cell in block['cells']:
                excel_cell = ws.cell(row=row_index, column=column_index, value=cell['value'] or None)
                excel_cell.alignment = center
                excel_cell.border = thin_border
                column_index += 1

            avg_cell = ws.cell(row=row_index, column=column_index, value=block['competency_grade'] or '-')
            avg_cell.alignment = center
            avg_cell.border = thin_border
            column_index += 1

        course_grade_cell = ws.cell(row=row_index, column=column_index, value=row['course_grade'] or '-')
        course_grade_cell.alignment = center
        course_grade_cell.border = thin_border

    for row in ws.iter_rows(min_row=1, max_row=max(len(rows) + 6, 6), min_col=1, max_col=last_column):
        for cell in row:
            cell.border = thin_border

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 100
    for row_index in range(7, 7 + len(rows)):
        ws.row_dimensions[row_index].height = 20

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    current_column = 3
    for block in competency_blocks:
        for _indicator in block['indicators']:
            ws.column_dimensions[get_column_letter(current_column)].width = 10
            current_column += 1
        ws.column_dimensions[get_column_letter(current_column)].width = 11
        current_column += 1
    ws.column_dimensions[get_column_letter(current_column)].width = 12

    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    grade_slug = slugify(
        selected_assignment.section.grade.name if selected_assignment.section else (selected_assignment.grade.name if selected_assignment.grade else 'sin_grado')
    ) or 'sin_grado'
    section_slug = slugify(selected_assignment.section.name) if selected_assignment.section else 'todas_secciones'
    period_slug = slugify(selected_period.name) or 'periodo'
    unit_slug = slugify(selected_unit.name) or 'unidad'
    course_slug = slugify(selected_assignment.course.name) or 'curso'
    filename = f"notas_{course_slug}_{grade_slug}_{section_slug}_{period_slug}_{unit_slug}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('admin', 'director', 'teacher')
def manage_competencies(request, assignment_id):
    assignment_qs = TeacherCourseAssignment.objects.select_related('teacher', 'course', 'section__grade', 'academic_year')
    assignment = assignment_qs.filter(id=assignment_id).first()
    if not assignment:
        course = Course.objects.filter(id=assignment_id).first()
        if course:
            assignment_qs = assignment_qs.filter(course=course)
            if request.user.role == 'teacher' and not request.user.is_superuser:
                assignment_qs = assignment_qs.filter(teacher=request.user)
            assignment = assignment_qs.order_by('-academic_year__year', 'section__grade__name', 'section__name', 'id').first()
            if assignment:
                messages.warning(
                    request,
                    "Ese enlace apuntaba al curso. Se abrió la asignacion correcta para gestionar las competencias.",
                )
    if not assignment:
        raise Http404("No TeacherCourseAssignment matches the given query.")
    
    # Security check: if teacher, must own this assignment.
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if assignment.teacher_id != request.user.id:
            messages.error(request, "No tienes permiso para gestionar esta seccion.")
            return redirect('teacher_dashboard')

    if request.method == 'POST':
        form = CompetencyForm(request.POST)
        if form.is_valid():
            competency = form.save(commit=False)
            competency.assignment = assignment
            last_order = (
                Competency.objects.filter(assignment=assignment)
                .aggregate(max_order=models.Max('order'))
                .get('max_order')
                or 0
            )
            competency.order = last_order + 1
            competency.save()
            messages.success(request, f"Competencia '{competency.name}' agregada.")
            return redirect('manage_competencies', assignment_id=assignment.id)
    else:
        form = CompetencyForm()

    competencies = Competency.objects.filter(assignment=assignment).prefetch_related('indicator_set')
    return render(request, 'academic/manage_competencies.html', {
        'assignment': assignment,
        'competencies': competencies,
        'form': form
    })


@role_required('admin', 'director', 'teacher')
def delete_competency(request, competency_id):
    competency = get_object_or_404(Competency, id=competency_id)
    assignment_id = competency.assignment_id
    
    # Security check
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if competency.assignment.teacher_id != request.user.id:
            return HttpResponse("Unauthorized", status=401)

    competency.delete()
    messages.success(request, "Competencia eliminada.")
    return redirect('manage_competencies', assignment_id=assignment_id)


@role_required('admin', 'director', 'teacher')
def manage_indicators(request, competency_id):
    competency = get_object_or_404(Competency, id=competency_id)
    assignment = competency.assignment
    period_id = request.POST.get('period') if request.method == 'POST' else request.GET.get('period')
    selected_period = Period.objects.filter(academic_year=assignment.academic_year, id=period_id).first()
    if not selected_period:
        selected_period = _preferred_period(assignment.academic_year)

    units = _ensure_assignment_units(assignment, selected_period)
    unit_id = request.POST.get('unit') if request.method == 'POST' else request.GET.get('unit')
    selected_unit = next((unit for unit in units if str(unit.id) == str(unit_id)), None) if unit_id else units[0]
    if not selected_unit and units:
        selected_unit = units[0]
    previous_unit = next((unit for unit in units if unit.order == selected_unit.order - 1), None) if selected_unit else None
    
    # Security check
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if assignment.teacher_id != request.user.id:
            messages.error(request, "No tienes permiso para gestionar esto.")
            return redirect('teacher_dashboard')

    if request.method == 'POST':
        if request.POST.get('replicate_previous_unit'):
            if not previous_unit:
                messages.error(request, 'No hay una unidad anterior para replicar.')
                return redirect(f"{request.path}?period={selected_period.id}&unit={selected_unit.id}")

            created = _replicate_indicators_to_unit(competency, previous_unit, selected_unit)
            messages.success(request, f'Se replicaron {created} indicadores a {selected_unit.name}.')
            return redirect(f"{request.path}?period={selected_period.id}&unit={selected_unit.id}")

        form = IndicatorForm(request.POST)
        if form.is_valid():
            indicator = form.save(commit=False)
            indicator.competency = competency
            indicator.unit = selected_unit
            last_order = (
                Indicator.objects.filter(competency=competency, unit=selected_unit)
                .aggregate(max_order=models.Max('order'))
                .get('max_order')
                or 0
            )
            indicator.order = last_order + 1
            indicator.save()
            messages.success(request, f"Indicador '{indicator.name}' agregado.")
            return redirect(f"{request.path}?period={selected_period.id}&unit={selected_unit.id}")
    else:
        form = IndicatorForm()

    indicators = Indicator.objects.filter(competency=competency, unit=selected_unit)
    return render(request, 'academic/manage_indicators.html', {
        'competency': competency,
        'assignment': assignment,
        'periods': Period.objects.filter(academic_year=assignment.academic_year).order_by('start_date', 'name'),
        'selected_period': selected_period,
        'units': units,
        'selected_unit': selected_unit,
        'previous_unit': previous_unit,
        'indicators': indicators,
        'form': form
    })


@role_required('admin', 'director', 'teacher')
def delete_indicator(request, indicator_id):
    indicator = get_object_or_404(Indicator, id=indicator_id)
    competency_id = indicator.competency_id
    
    # Security check
    if request.user.role == 'teacher' and not request.user.is_superuser:
        if indicator.competency.assignment.teacher_id != request.user.id:
            return HttpResponse("Unauthorized", status=401)

    indicator.delete()
    messages.success(request, "Indicador eliminado.")
    return redirect('manage_indicators', competency_id=competency_id)
@role_required('admin', 'director')
def auto_assign_poly_courses(request):
    from accounts.models import User
    
    # 1. Get all courses marked as poly_course
    poly_courses = Course.objects.filter(is_poly_course=True)
    if not poly_courses.exists():
        messages.warning(request, "No hay cursos marcados como 'Polidocencia'. Por favor, edita los cursos y marca esta opción.")
        return redirect('course_management')

    # 2. Get active academic year
    active_year = AcademicYear.objects.filter(is_active=True).order_by('-year').first()
    if not active_year:
        messages.error(request, "No hay un año académico activo.")
        return redirect('course_management')

    # 3. Get all sections
    sections = Section.objects.select_related('grade').all()
    count = 0
    assigned_info = []
    
    with transaction.atomic():
        for section in sections:
            grade_name = section.grade.name.lower()
            
            for course in poly_courses:
                should_assign = False
                name = course.name
                
                if "ciencia" in name.lower():
                    # Special Rule for Science: 4th to 6th grade
                    is_4_to_6 = any(x in grade_name for x in ["4", "5", "6"])
                    if is_4_to_6:
                        should_assign = True
                else:
                    # Others are for all grades
                    should_assign = True
                
                if should_assign:
                    # Find a teacher linked to this course
                    teacher = User.objects.filter(role='teacher', is_polyteacher=True, poly_course=course).first()
                    
                    if not teacher:
                        # If no poly teacher specifically linked to the FK, check M2M for retrocompat
                        teacher = User.objects.filter(role='teacher', is_polyteacher=True, teaching_courses=course).first()

                    if teacher:
                        obj, created = TeacherCourseAssignment.objects.get_or_create(
                            course=course,
                            section=section,
                            academic_year=active_year,
                            defaults={
                                'grade': section.grade,
                                'teacher': teacher,
                            }
                        )
                        if created:
                            count += 1
                    else:
                        assigned_info.append(f"No se pudo asignar '{name}' a {section} porque no hay un docente polidocente vinculado a este curso.")

    if count > 0:
        messages.success(request, f"Se han creado {count} asignaciones de cursos de polidocencia.")
    
    if assigned_info:
        for info in set(assigned_info): # Use set to avoid repeating same message for every section
            messages.warning(request, info)
            
    if count == 0 and not assigned_info:
        messages.info(request, "No se realizaron nuevas asignaciones. Es posible que ya estén configuradas.")

    return redirect('course_management')
