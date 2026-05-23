import calendar
import os
from collections import defaultdict
from datetime import date
from io import BytesIO

from django.db.models import Count
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from academic.models import AcademicYear
from core.student_ordering import order_queryset_by_student_name
from enrollment.models import Enrollment
from schools.models import School

from .models import AttendanceRecord


MONTH_NAMES = (
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
)
WEEKDAY_SHORT = ('Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do')
STATUS_MARKS = {
    'present': '\u2713',
    'absent': 'F',
    'tardy': 'T',
    'justified': 'J',
    'missing': '',
}
STATUS_LABELS = {
    'present': 'Asistencia',
    'absent': 'Falta',
    'tardy': 'Tardanza',
    'justified': 'Justificado',
    'missing': 'Sin registro',
}
STATUS_THEME = {
    'present': {'fill': 'DDF3E4', 'font': '256B45', 'pdf_fill': colors.HexColor('#dff2e3'), 'css': 'is-present'},
    'absent': {'fill': 'FBE2E1', 'font': '8C2D2A', 'pdf_fill': colors.HexColor('#fbe3e0'), 'css': 'is-absent'},
    'tardy': {'fill': 'FFF2CC', 'font': '8A6B00', 'pdf_fill': colors.HexColor('#fff3cf'), 'css': 'is-tardy'},
    'justified': {'fill': 'DCEAF7', 'font': '345C7C', 'pdf_fill': colors.HexColor('#ddebf7'), 'css': 'is-justified'},
    'missing': {'fill': 'F6F8FB', 'font': '8FA0AF', 'pdf_fill': colors.HexColor('#f5f7fa'), 'css': 'is-missing'},
}
HEADER_FILL = 'D9E2F3'
HEADER_ALT_FILL = 'EAF0F8'
SUBHEADER_FILL = 'EDF3F9'
SUMMARY_FILL = 'E8EFF7'
BORDER_COLOR = 'A8B7C7'


def _resolve_academic_year(selected_date):
    year_obj = AcademicYear.objects.filter(year=selected_date.year).first()
    if year_obj:
        return year_obj
    return AcademicYear.objects.filter(is_active=True).order_by('-year').first()


def _active_enrollments_for_section(section, selected_date, student_order='az'):
    filters = {
        'status': 'active',
        'section': section,
    }
    year_obj = _resolve_academic_year(selected_date)
    if year_obj:
        filters['academic_year'] = year_obj
    return order_queryset_by_student_name(
        Enrollment.objects.select_related('student').filter(**filters),
        prefix='student',
        student_order=student_order,
    )


def _person_name(user):
    if not user:
        return 'No asignado'
    full_name = user.get_full_name().strip()
    return full_name or user.username


def _smart_title(value):
    return ' '.join(part.capitalize() for part in (value or '').split())


def _format_student_name_for_report(student):
    first_tokens = [token for token in (student.first_name or '').split() if token]
    last_tokens = [token for token in (student.last_name or '').split() if token]

    if not first_tokens and not last_tokens:
        return ''
    if not last_tokens:
        return _smart_title(' '.join(first_tokens))
    if not first_tokens:
        return _smart_title(' '.join(last_tokens))

    if len(first_tokens) >= 2:
        surnames = ' '.join(first_tokens)
        names = ' '.join(last_tokens)
        return f'{_smart_title(surnames)}, {_smart_title(names)}'

    if len(last_tokens) >= 2:
        surnames = f"{first_tokens[0]} {last_tokens[0]}"
        names = ' '.join(last_tokens[1:])
        return f'{_smart_title(surnames)}, {_smart_title(names)}'

    return f'{_smart_title(" ".join(last_tokens))}, {_smart_title(" ".join(first_tokens))}'


def _build_month_days(selected_date):
    weeks_by_index = {}
    for week_index, week_dates in enumerate(
        calendar.Calendar(firstweekday=0).monthdatescalendar(selected_date.year, selected_date.month),
        start=1,
    ):
        month_dates = [current for current in week_dates if current.month == selected_date.month and current.weekday() < 5]
        if not month_dates:
            continue
        weeks_by_index[week_index] = {
            'index': week_index,
            'label': f'Semana {week_index}',
            'colspan': 0,
        }

    days = []
    for day_number in range(1, calendar.monthrange(selected_date.year, selected_date.month)[1] + 1):
        current = date(selected_date.year, selected_date.month, day_number)
        if current.weekday() >= 5:
            continue
        week_index = next(
            index
            for index, week_dates in enumerate(
                calendar.Calendar(firstweekday=0).monthdatescalendar(selected_date.year, selected_date.month),
                start=1,
            )
            if current in week_dates
        )
        weeks_by_index[week_index]['colspan'] += 1
        starts_week = not days or days[-1]['week_index'] != week_index
        days.append({
            'date': current,
            'day': day_number,
            'weekday_short': WEEKDAY_SHORT[current.weekday()],
            'week_index': week_index,
            'starts_week': starts_week,
        })

    weeks = [weeks_by_index[index] for index in sorted(weeks_by_index)]
    return days, weeks


def build_attendance_report_data(section, selected_date, student_order='az', request_user=None):
    del request_user
    month_start = selected_date.replace(day=1)
    month_end = selected_date.replace(day=calendar.monthrange(selected_date.year, selected_date.month)[1])
    enrollments = list(_active_enrollments_for_section(section, selected_date, student_order=student_order))
    records = AttendanceRecord.objects.select_related(
        'enrollment__student',
    ).filter(
        enrollment__in=enrollments,
        date__range=(month_start, month_end),
        date__week_day__in=[2, 3, 4, 5, 6],
    )
    records_by_key = {(record.enrollment_id, record.date): record for record in records}
    days, weeks = _build_month_days(selected_date)

    school = School.objects.order_by('id').first()
    logo_path = None
    logo_url = ''
    if school and school.logo:
        logo_url = school.logo.url
        try:
            candidate = school.logo.path
        except (NotImplementedError, ValueError):
            candidate = None
        if candidate and os.path.exists(candidate):
            logo_path = candidate

    totals = defaultdict(int)
    rows = []
    for index, enrollment in enumerate(enrollments, start=1):
        counts = {'present': 0, 'absent': 0, 'tardy': 0, 'justified': 0, 'missing': 0}
        cells = []
        for day in days:
            record = records_by_key.get((enrollment.id, day['date']))
            status = record.status if record else 'missing'
            counts[status] += 1
            cells.append({
                'status': status,
                'short': STATUS_MARKS[status],
                'label': STATUS_LABELS[status],
                'css_class': STATUS_THEME[status]['css'],
                'starts_week': day['starts_week'],
            })

        recorded_days = counts['present'] + counts['absent'] + counts['tardy'] + counts['justified']
        attended_days = counts['present'] + counts['tardy'] + counts['justified']
        attendance_pct = round((attended_days / recorded_days) * 100, 2) if recorded_days else 0

        rows.append({
            'number': index,
            'student_name': _format_student_name_for_report(enrollment.student),
            'cells': cells,
            'counts': counts,
            'attendance_pct': attendance_pct,
        })
        for key in counts:
            totals[key] += counts[key]

    total_recorded_marks = totals['present'] + totals['absent'] + totals['tardy'] + totals['justified']
    total_attended_marks = totals['present'] + totals['tardy'] + totals['justified']
    overall_pct = round((total_attended_marks / total_recorded_marks) * 100, 2) if total_recorded_marks else 0

    return {
        'school_name': school.name if school else 'Institucion educativa',
        'logo_path': logo_path,
        'logo_url': logo_url,
        'teacher_name': _person_name(section.tutor_teacher),
        'grade_section': f"{section.grade.name} {section.name}",
        'month_label': f"{MONTH_NAMES[selected_date.month - 1]} {selected_date.year}",
        'generated_at': timezone.localtime(),
        'days': days,
        'weeks': weeks,
        'rows': rows,
        'legend': [
            {'short': '\u2713', 'label': 'Asistencia', 'css_class': STATUS_THEME['present']['css']},
            {'short': 'F', 'label': 'Falta', 'css_class': STATUS_THEME['absent']['css']},
            {'short': 'T', 'label': 'Tardanza', 'css_class': STATUS_THEME['tardy']['css']},
            {'short': 'J', 'label': 'Justificado', 'css_class': STATUS_THEME['justified']['css']},
        ],
        'totals': dict(totals),
        'student_count': len(rows),
        'days_with_records': records.values('date').annotate(total=Count('id')).count(),
        'overall_attendance_pct': overall_pct,
    }


def _border(style='thin'):
    return Side(style=style, color=BORDER_COLOR)


def _apply_border(cell, left='thin', right='thin', top='thin', bottom='thin'):
    cell.border = Border(
        left=_border(left),
        right=_border(right),
        top=_border(top),
        bottom=_border(bottom),
    )


def _apply_meta_block(ws, row, start_col, end_col, text):
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    ws.merge_cells(f'{start_letter}{row}:{end_letter}{row}')
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.fill = PatternFill('solid', fgColor=SUBHEADER_FILL)
    cell.font = Font(name='Calibri', size=10, bold=True, color='25364A')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    _apply_border(cell)


def build_attendance_workbook(report_data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Asistencia'
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.print_options.horizontalCentered = True
    worksheet.freeze_panes = 'C9'

    total_day_columns = len(report_data['days'])
    day_start_col = 3
    day_end_col = day_start_col + total_day_columns - 1
    percent_col = day_end_col + 1
    absent_col = day_end_col + 2
    tardy_col = day_end_col + 3
    justified_col = day_end_col + 4
    table_end_col = justified_col
    table_end_letter = get_column_letter(table_end_col)

    worksheet.column_dimensions['A'].width = 6
    worksheet.column_dimensions['B'].width = 34
    for column in range(day_start_col, day_end_col + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 4.2
    worksheet.column_dimensions[get_column_letter(percent_col)].width = 12
    worksheet.column_dimensions[get_column_letter(absent_col)].width = 9
    worksheet.column_dimensions[get_column_letter(tardy_col)].width = 9
    worksheet.column_dimensions[get_column_letter(justified_col)].width = 9

    if report_data['logo_path']:
        image = XLImage(report_data['logo_path'])
        image.width = 70
        image.height = 70
        worksheet.add_image(image, 'A1')

    worksheet.merge_cells(f'C1:{table_end_letter}1')
    worksheet['C1'] = report_data['school_name']
    worksheet['C1'].font = Font(name='Calibri', size=16, bold=True, color='203040')
    worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['C1'].fill = PatternFill('solid', fgColor=HEADER_FILL)
    _apply_border(worksheet['C1'])

    worksheet.merge_cells(f'C2:{table_end_letter}2')
    worksheet['C2'] = 'Reporte mensual de asistencia'
    worksheet['C2'].font = Font(name='Calibri', size=12, bold=True, color='3F556B')
    worksheet['C2'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['C2'].fill = PatternFill('solid', fgColor=HEADER_FILL)
    _apply_border(worksheet['C2'])

    split_a = max(8, table_end_col // 3)
    split_b = max(split_a + 1, (table_end_col * 2) // 3)
    _apply_meta_block(worksheet, 3, 1, split_a, f'Docente: {report_data["teacher_name"]}')
    _apply_meta_block(worksheet, 3, split_a + 1, split_b, f'Grado y seccion: {report_data["grade_section"]}')
    _apply_meta_block(worksheet, 3, split_b + 1, table_end_col, f'Mes y ano: {report_data["month_label"]}')
    _apply_meta_block(
        worksheet,
        4,
        1,
        table_end_col,
        f'Leyenda: \u2713 Asistencia | F Falta | T Tardanza | J Justificado | Generado: {report_data["generated_at"].strftime("%d/%m/%Y %H:%M")}',
    )

    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 22
    worksheet.row_dimensions[4].height = 20

    week_row = 6
    day_row = 7
    weekday_row = 8
    data_start_row = 9
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    header_font = Font(name='Calibri', size=9, bold=True, color='30475E')

    worksheet.merge_cells(start_row=week_row, start_column=1, end_row=weekday_row, end_column=1)
    worksheet.merge_cells(start_row=week_row, start_column=2, end_row=weekday_row, end_column=2)
    worksheet.cell(row=week_row, column=1, value='Nro')
    worksheet.cell(row=week_row, column=2, value='Apellidos y nombres')

    for column in (1, 2):
        cell = worksheet.cell(row=week_row, column=column)
        cell.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cell.font = Font(name='Calibri', size=10, bold=True, color='24384B')
        cell.alignment = center if column == 1 else left
        _apply_border(cell, left='medium' if column == 1 else 'thin', top='medium', bottom='medium')

    current_col = day_start_col
    for week in report_data['weeks']:
        start_col = current_col
        end_col = current_col + week['colspan'] - 1
        worksheet.merge_cells(start_row=week_row, start_column=start_col, end_row=week_row, end_column=end_col)
        cell = worksheet.cell(row=week_row, column=start_col, value=week['label'])
        cell.fill = PatternFill('solid', fgColor=HEADER_FILL if week['index'] % 2 else HEADER_ALT_FILL)
        cell.font = header_font
        cell.alignment = center
        _apply_border(cell, top='medium', bottom='thin')
        current_col = end_col + 1

    worksheet.merge_cells(start_row=week_row, start_column=percent_col, end_row=week_row, end_column=justified_col)
    summary_group = worksheet.cell(row=week_row, column=percent_col, value='Resumen')
    summary_group.fill = PatternFill('solid', fgColor=HEADER_FILL)
    summary_group.font = header_font
    summary_group.alignment = center
    _apply_border(summary_group, top='medium', bottom='thin')

    for offset, day in enumerate(report_data['days']):
        column = day_start_col + offset
        for row_number, value in ((day_row, day['day']), (weekday_row, day['weekday_short'])):
            cell = worksheet.cell(row=row_number, column=column, value=value)
            cell.fill = PatternFill('solid', fgColor=SUBHEADER_FILL)
            cell.font = header_font
            cell.alignment = center
            _apply_border(cell, left='medium' if day['starts_week'] else 'thin')

    for column, label in (
        (percent_col, '% Asist.'),
        (absent_col, 'Faltas'),
        (tardy_col, 'Tard.'),
        (justified_col, 'Justif.'),
    ):
        worksheet.merge_cells(start_row=day_row, start_column=column, end_row=weekday_row, end_column=column)
        cell = worksheet.cell(row=day_row, column=column, value=label)
        cell.fill = PatternFill('solid', fgColor=SUMMARY_FILL)
        cell.font = header_font
        cell.alignment = center
        _apply_border(cell)

    day_range_start_letter = get_column_letter(day_start_col)
    day_range_end_letter = get_column_letter(day_end_col)
    present_mark = STATUS_MARKS['present']

    for row_index, row in enumerate(report_data['rows'], start=data_start_row):
        worksheet.row_dimensions[row_index].height = 21
        number_cell = worksheet.cell(row=row_index, column=1, value=row['number'])
        student_cell = worksheet.cell(row=row_index, column=2, value=row['student_name'])
        number_cell.font = Font(name='Calibri', size=10, color='25364A')
        number_cell.alignment = center
        student_cell.font = Font(name='Calibri', size=10, color='25364A')
        student_cell.alignment = left
        _apply_border(number_cell)
        _apply_border(student_cell)

        for offset, cell_data in enumerate(row['cells']):
            column = day_start_col + offset
            cell = worksheet.cell(row=row_index, column=column, value=cell_data['short'])
            cell.font = Font(name='Calibri', size=10, bold=True, color=STATUS_THEME[cell_data['status']]['font'])
            cell.alignment = center
            cell.fill = PatternFill('solid', fgColor=STATUS_THEME[cell_data['status']]['fill'])
            _apply_border(cell, left='medium' if cell_data['starts_week'] else 'thin')

        excel_day_range = f'{day_range_start_letter}{row_index}:{day_range_end_letter}{row_index}'
        worksheet.cell(
            row=row_index,
            column=percent_col,
            value=(
                f'=IF(COUNTIF({excel_day_range},"<>")=0,0,'
                f'(COUNTIF({excel_day_range},"{present_mark}")+COUNTIF({excel_day_range},"T")+COUNTIF({excel_day_range},"J"))/'
                f'COUNTIF({excel_day_range},"<>"))'
            ),
        )
        worksheet.cell(row=row_index, column=absent_col, value=f'=COUNTIF({excel_day_range},"F")')
        worksheet.cell(row=row_index, column=tardy_col, value=f'=COUNTIF({excel_day_range},"T")')
        worksheet.cell(row=row_index, column=justified_col, value=f'=COUNTIF({excel_day_range},"J")')

        for column in (percent_col, absent_col, tardy_col, justified_col):
            cell = worksheet.cell(row=row_index, column=column)
            cell.font = Font(name='Calibri', size=10, color='25364A')
            cell.alignment = center
            cell.fill = PatternFill('solid', fgColor='F8FBFD')
            _apply_border(cell)
        worksheet.cell(row=row_index, column=percent_col).number_format = '0.00%'

    totals_row = data_start_row + len(report_data['rows'])
    worksheet.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=2)
    totals_label = worksheet.cell(row=totals_row, column=1, value='Totales del mes')
    totals_label.fill = PatternFill('solid', fgColor=HEADER_FILL)
    totals_label.font = Font(name='Calibri', size=10, bold=True, color='203040')
    totals_label.alignment = left
    _apply_border(totals_label, left='medium', top='medium', bottom='medium')

    for column in range(day_start_col, day_end_col + 1):
        cell = worksheet.cell(row=totals_row, column=column, value='')
        cell.fill = PatternFill('solid', fgColor='F8FBFD')
        _apply_border(cell, top='medium', bottom='medium')

    if report_data['rows']:
        start_row = data_start_row
        end_row = totals_row - 1
        worksheet.cell(row=totals_row, column=percent_col, value=f'=AVERAGE({get_column_letter(percent_col)}{start_row}:{get_column_letter(percent_col)}{end_row})')
        worksheet.cell(row=totals_row, column=absent_col, value=f'=SUM({get_column_letter(absent_col)}{start_row}:{get_column_letter(absent_col)}{end_row})')
        worksheet.cell(row=totals_row, column=tardy_col, value=f'=SUM({get_column_letter(tardy_col)}{start_row}:{get_column_letter(tardy_col)}{end_row})')
        worksheet.cell(row=totals_row, column=justified_col, value=f'=SUM({get_column_letter(justified_col)}{start_row}:{get_column_letter(justified_col)}{end_row})')
        worksheet.cell(row=totals_row, column=percent_col).number_format = '0.00%'

    for column in (percent_col, absent_col, tardy_col, justified_col):
        cell = worksheet.cell(row=totals_row, column=column)
        cell.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cell.font = Font(name='Calibri', size=10, bold=True, color='203040')
        cell.alignment = center
        _apply_border(cell, top='medium', bottom='medium')

    worksheet.print_title_rows = '1:8'
    return workbook


def workbook_to_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_attendance_pdf(report_data):
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('AttendanceTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#24384B'), alignment=TA_CENTER, spaceAfter=6)
    meta_style = ParagraphStyle('AttendanceMeta', parent=styles['BodyText'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#42586D'), alignment=TA_LEFT)
    legend_style = ParagraphStyle('AttendanceLegend', parent=styles['BodyText'], fontName='Helvetica-Bold', fontSize=7, leading=9, textColor=colors.HexColor('#314659'))

    story = []
    logo = Image(report_data['logo_path'], width=18 * mm, height=18 * mm) if report_data['logo_path'] else None
    info_lines = [
        Paragraph(report_data['school_name'], title_style),
        Paragraph('Reporte mensual de asistencia', meta_style),
        Paragraph(
            f'<b>Docente:</b> {report_data["teacher_name"]} &nbsp;&nbsp; <b>Grado y seccion:</b> {report_data["grade_section"]} &nbsp;&nbsp; <b>Mes y ano:</b> {report_data["month_label"]}',
            meta_style,
        ),
    ]
    header_table = Table([[logo, info_lines]] if logo else [[info_lines]], colWidths=[22 * mm, 245 * mm] if logo else [267 * mm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph('<b>Leyenda:</b> &#10003; Asistencia &nbsp;&nbsp; F Falta &nbsp;&nbsp; T Tardanza &nbsp;&nbsp; J Justificado', legend_style))
    story.append(Spacer(1, 3 * mm))

    week_row = ['Nro', 'Apellidos y nombres']
    day_row = ['', '']
    weekday_row = ['', '']
    spans = [('SPAN', (0, 0), (0, 2)), ('SPAN', (1, 0), (1, 2))]

    current_col = 2
    for week in report_data['weeks']:
        week_row.extend([week['label']] + [''] * (week['colspan'] - 1))
        spans.append(('SPAN', (current_col, 0), (current_col + week['colspan'] - 1, 0)))
        current_col += week['colspan']

    day_row.extend([day['day'] for day in report_data['days']])
    weekday_row.extend([day['weekday_short'] for day in report_data['days']])

    summary_start = 2 + len(report_data['days'])
    week_row.extend(['Resumen', '', '', ''])
    spans.append(('SPAN', (summary_start, 0), (summary_start + 3, 0)))
    day_row.extend(['% Asist.', 'Faltas', 'Tard.', 'Justif.'])
    weekday_row.extend(['', '', '', ''])
    for column in range(summary_start, summary_start + 4):
        spans.append(('SPAN', (column, 1), (column, 2)))

    table_data = [week_row, day_row, weekday_row]
    for row in report_data['rows']:
        data_row = [row['number'], row['student_name']]
        data_row.extend([cell['short'] for cell in row['cells']])
        data_row.extend([f'{row["attendance_pct"]:.2f}%', row['counts']['absent'], row['counts']['tardy'], row['counts']['justified']])
        table_data.append(data_row)

    if report_data['rows']:
        totals_row = ['Totales del mes', '']
        totals_row.extend(['' for _ in report_data['days']])
        totals_row.extend([f'{report_data["overall_attendance_pct"]:.2f}%', report_data['totals']['absent'], report_data['totals']['tardy'], report_data['totals']['justified']])
        table_data.append(totals_row)
        spans.append(('SPAN', (0, len(table_data) - 1), (1, len(table_data) - 1)))

    column_widths = [8 * mm, 45 * mm]
    column_widths.extend([7 * mm for _ in report_data['days']])
    column_widths.extend([12 * mm, 10 * mm, 10 * mm, 10 * mm])

    table = Table(table_data, colWidths=column_widths, repeatRows=3)
    style_commands = [
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#a8b7c7')),
        ('BACKGROUND', (0, 0), (1, 2), colors.HexColor('#d9e2f3')),
        ('BACKGROUND', (summary_start, 0), (-1, 2), colors.HexColor('#e8eff7')),
        ('BACKGROUND', (0, 1), (1 + len(report_data['days']), 2), colors.HexColor('#edf3f9')),
        ('FONTNAME', (0, 0), (-1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 2), colors.HexColor('#2f455b')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 2), 6.8),
        ('FONTSIZE', (0, 3), (-1, -1), 6.4),
        ('LEADING', (0, 0), (-1, -1), 7.2),
        ('ALIGN', (1, 3), (1, -1), 'LEFT'),
        ('LEFTPADDING', (1, 3), (1, -1), 3),
        ('RIGHTPADDING', (1, 3), (1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    style_commands.extend(spans)

    for row_index, row in enumerate(report_data['rows'], start=3):
        for cell_index, cell_data in enumerate(row['cells'], start=2):
            style_commands.append(('BACKGROUND', (cell_index, row_index), (cell_index, row_index), STATUS_THEME[cell_data['status']]['pdf_fill']))
            style_commands.append(('TEXTCOLOR', (cell_index, row_index), (cell_index, row_index), colors.HexColor(f'#{STATUS_THEME[cell_data["status"]]["font"]}')))
        if row_index % 2 == 0:
            style_commands.append(('BACKGROUND', (0, row_index), (1, row_index), colors.HexColor('#f8fbfd')))

    if report_data['rows']:
        totals_index = len(table_data) - 1
        style_commands.extend([
            ('BACKGROUND', (0, totals_index), (-1, totals_index), colors.HexColor('#d9e2f3')),
            ('FONTNAME', (0, totals_index), (-1, totals_index), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, totals_index), (-1, totals_index), colors.HexColor('#24384B')),
        ])

    table.setStyle(TableStyle(style_commands))
    story.append(table)
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(f'Generado el {report_data["generated_at"].strftime("%d/%m/%Y %H:%M")} | Alumnos: {report_data["student_count"]} | Dias con registros: {report_data["days_with_records"]}', meta_style))
    document.build(story)
    return buffer.getvalue()
