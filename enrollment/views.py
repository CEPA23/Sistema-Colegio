from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q

from accounts.decorators import role_required
from academic.models import Grade, Section
from students.models import Student

from .forms import EnrollmentForm, StudentBulkImportForm
from .models import Enrollment


@role_required('admin', 'director', 'secretary')
def enrollment_dashboard(request):
    context = {
        'total_enrollments': Enrollment.objects.count(),
        'active_enrollments': Enrollment.objects.filter(status='active').count(),
        'retired_enrollments': Enrollment.objects.filter(status='retired').count(),
        'transferred_enrollments': Enrollment.objects.filter(status='transferred').count(),
    }
    return render(request, 'enrollment/enrollment_dashboard.html', context)


@role_required('admin', 'director', 'secretary')
def enrollment_list(request):
    enrollments = Enrollment.objects.select_related(
        'student',
        'section__grade',
        'academic_year'
    )

    q = (request.GET.get('q') or '').strip()
    grade_id = request.GET.get('grade')
    section_id = request.GET.get('section')

    if grade_id and str(grade_id).isdigit():
        enrollments = enrollments.filter(section__grade_id=int(grade_id))

    if section_id and str(section_id).isdigit():
        enrollments = enrollments.filter(section_id=int(section_id))

    if q:
        enrollments = enrollments.filter(
            Q(student__first_name__icontains=q)
            | Q(student__last_name__icontains=q)
            | Q(student__dni__icontains=q)
        )

    enrollments = enrollments.order_by('-enrolled_at')
    filtered_total = enrollments.count()
    grades = Grade.objects.all().order_by('name')
    sections = Section.objects.select_related('grade').order_by('grade__name', 'name')
    if grade_id and str(grade_id).isdigit():
        sections = sections.filter(grade_id=int(grade_id))
    return render(request, 'enrollment/enrollment_list.html', {
        'enrollments': enrollments,
        'grades': grades,
        'sections': sections,
        'selected_grade': int(grade_id) if grade_id and str(grade_id).isdigit() else None,
        'selected_section': int(section_id) if section_id and str(section_id).isdigit() else None,
        'q': q,
        'filtered_total': filtered_total,
    })


def _normalize_header(value):
    import re
    import unicodedata

    raw = (str(value).strip().lower() if value is not None else '')
    raw = ''.join(
        c for c in unicodedata.normalize('NFKD', raw)
        if not unicodedata.combining(c)
    )
    raw = re.sub(r'[^a-z0-9]+', '', raw)
    return raw


def _coerce_str(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _parse_date(value):
    from datetime import datetime, date

    if value is None or value == '':
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    raw = _coerce_str(value)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _split_full_name(value):
    raw = _coerce_str(value)
    if not raw:
        return None, None

    if ',' in raw:
        left, right = raw.split(',', 1)
        last_name = left.strip()
        first_name = right.strip()
        if first_name and last_name:
            return first_name, last_name

    parts = [p for p in raw.split() if p.strip()]
    if len(parts) < 2:
        return None, None
    if len(parts) == 2:
        return parts[0], parts[1]

    first_name = ' '.join(parts[:-2]).strip()
    last_name = ' '.join(parts[-2:]).strip()
    if not first_name:
        first_name = parts[0]
        last_name = ' '.join(parts[1:])
    return first_name, last_name


def _build_fallback_name(row_num, raw_name):
    first_name, last_name = _split_full_name(raw_name)
    if first_name and last_name:
        return first_name, last_name, None
    if first_name and not last_name:
        return first_name, 'SIN APELLIDO', f"Fila {row_num}: faltaba apellido; se completo como 'SIN APELLIDO'."
    if raw_name:
        normalized = _coerce_str(raw_name)
        if normalized:
            return normalized, 'SIN APELLIDO', f"Fila {row_num}: el nombre venia incompleto; se completo como 'SIN APELLIDO'."
    return 'ALUMNO', f'IMPORTADO {row_num}', f"Fila {row_num}: faltaba el nombre; se completo con un nombre temporal."


def _generate_temp_dni(existing_dnis):
    candidate = 90000000
    while f"{candidate:08d}" in existing_dnis:
        candidate += 1
    return f"{candidate:08d}"


def _import_duplicate_key(academic_year_id, dni, first_name, last_name, grade_id, section_id, use_dni=True):
    if use_dni and dni and dni.isdigit() and len(dni) == 8:
        return ('dni', academic_year_id, dni)
    return (
        'identity',
        academic_year_id,
        _normalize_header(first_name),
        _normalize_header(last_name),
        grade_id,
        section_id,
    )


@role_required('admin', 'director', 'secretary')
def enrollment_import_template(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        messages.error(request, "Falta instalar openpyxl para generar la plantilla Excel.")
        return redirect('enrollment_list')

    headers = [
        'Estudiante',
        'DNI',
        'Grado',
        'Seccion',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append([
        'Juan Perez',
        '12345678',
        '1',
        'A',
    ])

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_alumnos.xlsx"'
    return response


@role_required('admin', 'director', 'secretary')
def enrollment_import_students(request):
    results = None
    row_errors = []
    row_warnings = []
    duplicate_rows = []

    if request.method == 'POST':
        form = StudentBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                from openpyxl import load_workbook
            except ImportError:
                messages.error(request, "Falta instalar openpyxl para importar Excel.")
                return redirect('enrollment_import_students')

            uploaded = form.cleaned_data['file']
            try:
                wb = load_workbook(uploaded, data_only=True)
            except Exception:
                messages.error(request, "No se pudo leer el Excel. Asegurate que sea un .xlsx valido.")
                return redirect('enrollment_import_students')

            ws = wb.active
            header_row = [
                _normalize_header(ws.cell(row=1, column=c).value)
                for c in range(1, ws.max_column + 1)
            ]
            header_index = {name: idx for idx, name in enumerate(header_row, start=1) if name}

            aliases = {
                'student_name': {'estudiante', 'alumno', 'student', 'nombreyapellido', 'nombresyapellidos'},
                'dni': {'dni', 'documento', 'documentodeidentidad'},
                'grade': {'grado', 'grade'},
                'section': {'seccion', 'seccionaula', 'section'},
                'father_name': {'nombrepadre', 'padre', 'fathername', 'father_name'},
                'father_phone': {'telefonopadre', 'celularpadre', 'fatherphone', 'father_phone'},
                'mother_name': {'nombremadre', 'madre', 'mothername', 'mother_name'},
                'mother_phone': {'telefonomadre', 'celularmadre', 'motherphone', 'mother_phone'},
            }

            def _col_for(field_key):
                for candidate in aliases[field_key]:
                    if candidate in header_index:
                        return header_index[candidate]
                return None

            columns = {key: _col_for(key) for key in aliases.keys()}
            if not columns.get('grade') and not columns.get('section'):
                messages.error(
                    request,
                    "Plantilla invalida: debe incluir al menos la columna Grado o Seccion para poder matricular."
                )
                return redirect('enrollment_import_students')

            academic_year = form.cleaned_data['academic_year']
            status = form.cleaned_data['status']

            grade_number_map = {}
            grade_slug_map = {}
            for grade in Grade.objects.all():
                grade_slug_map[_normalize_header(grade.name)] = grade
                import re

                match = re.search(r'\d+', grade.name or '')
                if match:
                    number = int(match.group(0))
                    grade_number_map.setdefault(number, []).append(grade)

            section_map = {}
            for sec in Section.objects.select_related('grade').all():
                section_map.setdefault(sec.grade_id, {})[_normalize_header(sec.name)] = sec

            created_students = 0
            existing_students = 0
            created_enrollments = 0
            skipped_enrollments = 0
            updated_students = 0
            duplicate_count = 0
            generated_dnis = {
                dni for dni in Student.objects.values_list('dni', flat=True)
                if isinstance(dni, str) and dni.isdigit() and len(dni) == 8
            }
            auto_generated_dnis = set()
            seen_import_keys = {}

            for row_num in range(2, ws.max_row + 1):
                raw_values = {
                    key: ws.cell(row=row_num, column=col).value if col else None
                    for key, col in columns.items()
                }
                if not any(v not in (None, '') for v in raw_values.values()):
                    continue

                first_name, last_name, name_warning = _build_fallback_name(row_num, raw_values.get('student_name'))
                if name_warning:
                    row_warnings.append(name_warning)

                dni = _coerce_str(raw_values['dni']).replace(' ', '').replace('-', '')
                using_real_dni = True
                if dni.isdigit():
                    dni = dni.zfill(8)
                if not dni.isdigit() or len(dni) != 8:
                    dni = _generate_temp_dni(generated_dnis)
                    generated_dnis.add(dni)
                    auto_generated_dnis.add(dni)
                    using_real_dni = False
                    row_warnings.append(
                        f"Fila {row_num}: DNI faltante o invalido; se genero un DNI temporal ({dni})."
                    )

                grade_name = _coerce_str(raw_values['grade'])
                section_name = _coerce_str(raw_values['section'])

                father_name = _coerce_str(raw_values.get('father_name'))
                father_phone = _coerce_str(raw_values.get('father_phone'))
                mother_name = _coerce_str(raw_values.get('mother_name'))
                mother_phone = _coerce_str(raw_values.get('mother_phone'))

                grade = None
                normalized_grade = _normalize_header(grade_name)
                if grade_name:
                    grade = grade_slug_map.get(normalized_grade) or Grade.objects.filter(name__iexact=grade_name).first()
                if grade is None:
                    import re

                    match = re.search(r'\d+', grade_name or '')
                    if match:
                        number = int(match.group(0))
                        candidates = grade_number_map.get(number, [])
                        if len(candidates) == 1:
                            grade = candidates[0]
                        elif len(candidates) > 1:
                            grade = (
                                next((g for g in candidates if _normalize_header(g.name) == normalized_grade), None)
                                or candidates[0]
                            )

                section = None
                if section_name:
                    if grade is not None:
                        section = (
                            section_map.get(grade.id, {}).get(_normalize_header(section_name))
                            or Section.objects.filter(grade=grade, name__iexact=section_name).first()
                        )
                    else:
                        matching_sections = list(Section.objects.filter(name__iexact=section_name).select_related('grade'))
                        if len(matching_sections) == 1:
                            section = matching_sections[0]
                            grade = section.grade
                            row_warnings.append(
                                f"Fila {row_num}: faltaba grado; se dedujo desde la seccion ({section.grade.name} {section.name})."
                            )
                elif grade is not None:
                    candidates = list(section_map.get(grade.id, {}).values())
                    if len(candidates) == 1:
                        section = candidates[0]
                        row_warnings.append(
                            f"Fila {row_num}: faltaba seccion; se uso la unica seccion del grado ({section.name})."
                        )

                if grade is None and section is None:
                    row_errors.append(
                        f"Fila {row_num}: falta grado/seccion o no se pudieron deducir para matricular al alumno."
                    )
                    continue

                if grade is None:
                    row_errors.append(
                        f"Fila {row_num}: grado no encontrado ({grade_name}). Ejemplo valido: '1' o '1 grado'."
                    )
                    continue

                if section is None:
                    if section_name:
                        row_errors.append(f"Fila {row_num}: seccion no encontrada para el grado ({grade_name} - {section_name}).")
                    else:
                        row_errors.append(f"Fila {row_num}: falta seccion y el grado ({grade_name}) tiene mas de una seccion.")
                    continue

                duplicate_key = _import_duplicate_key(
                    academic_year.id,
                    dni,
                    first_name,
                    last_name,
                    grade.id,
                    section.id,
                    use_dni=using_real_dni and dni not in auto_generated_dnis,
                )
                previous_row = seen_import_keys.get(duplicate_key)
                if previous_row is not None:
                    duplicate_count += 1
                    duplicate_rows.append(
                        f"Fila {row_num}: repetida con la fila {previous_row}; se omitio el duplicado."
                    )
                    continue
                seen_import_keys[duplicate_key] = row_num

                try:
                    with transaction.atomic():
                        student = Student.objects.filter(dni=dni).first()
                        if student is None:
                            student = Student.objects.create(
                                dni=dni,
                                first_name=first_name,
                                last_name=last_name,
                                father_name=father_name,
                                father_phone=father_phone,
                                mother_name=mother_name,
                                mother_phone=mother_phone,
                            )
                            created_students += 1
                        else:
                            existing_students += 1
                            update_fields = []
                            for field_name, field_value in (
                                ('father_name', father_name),
                                ('father_phone', father_phone),
                                ('mother_name', mother_name),
                                ('mother_phone', mother_phone),
                            ):
                                current = getattr(student, field_name)
                                if current in ('', None) and field_value not in ('', None):
                                    setattr(student, field_name, field_value)
                                    update_fields.append(field_name)
                            if update_fields:
                                student.save(update_fields=update_fields)
                                updated_students += 1

                        if Enrollment.objects.filter(student=student, academic_year=academic_year).exists():
                            skipped_enrollments += 1
                            row_warnings.append(
                                f"Fila {row_num}: el alumno ya tenia matricula en el anio {academic_year.year}; se omitio."
                            )
                        else:
                            Enrollment.objects.create(
                                student=student,
                                academic_year=academic_year,
                                section=section,
                                status=status,
                            )
                            created_enrollments += 1
                except Exception as exc:
                    row_errors.append(f"Fila {row_num}: error al guardar ({exc}).")

            results = {
                'created_students': created_students,
                'existing_students': existing_students,
                'updated_students': updated_students,
                'created_enrollments': created_enrollments,
                'skipped_enrollments': skipped_enrollments,
                'duplicate_count': duplicate_count,
                'error_count': len(row_errors),
                'warning_count': len(row_warnings),
            }

            messages.success(
                request,
                (
                    f"Importacion finalizada. "
                    f"Alumnos nuevos: {created_students}. "
                    f"Alumnos existentes: {existing_students} (actualizados: {updated_students}). "
                    f"Matriculas creadas: {created_enrollments}. "
                    f"Matriculas omitidas (ya existian en el anio): {skipped_enrollments}."
                )
            )
            if row_errors:
                messages.warning(request, f"Se encontraron {len(row_errors)} fila(s) con errores. Revisa el detalle.")
            if row_warnings:
                messages.info(request, f"Se completaron automaticamente {len(row_warnings)} dato(s) faltante(s).")
            if duplicate_rows:
                messages.warning(request, f"Se detectaron {len(duplicate_rows)} fila(s) repetida(s) y se omitieron.")
    else:
        form = StudentBulkImportForm()

    return render(request, 'enrollment/enrollment_import_students.html', {
        'form': form,
        'results': results,
        'row_errors': row_errors[:50],
        'row_warnings': row_warnings[:50],
        'duplicate_rows': duplicate_rows[:50],
    })


@role_required('admin', 'director', 'secretary')
def enrollment_create(request):
    messages.info(request, "La matricula nueva se registra desde Pagos > Registrar pago.")
    return redirect('payment_create')


@role_required('admin', 'director', 'secretary')
def enrollment_renew(request, enrollment_id):
    previous = get_object_or_404(Enrollment, id=enrollment_id)

    if request.method == 'POST':
        form = EnrollmentForm(request.POST)
        if form.is_valid():
            renewed = form.save(commit=False)
            renewed.student = previous.student
            renewed.status = 'active'
            renewed.save()
            messages.success(request, f"Matricula renovada para {renewed.student}.")
            return redirect('enrollment_list')
    else:
        form = EnrollmentForm(initial={
            'student': previous.student,
            'academic_year': previous.academic_year,
            'section': previous.section,
            'status': 'active',
        })
        form.fields['student'].disabled = True

    return render(request, 'enrollment/enrollment_renew_form.html', {'form': form, 'previous': previous})


@role_required('admin', 'director', 'secretary', 'teacher', 'parent')
def enrollment_history(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    history = Enrollment.objects.select_related(
        'section__grade',
        'academic_year'
    ).filter(student=student).order_by('-enrolled_at')
    return render(request, 'enrollment/enrollment_history.html', {'student': student, 'history': history})


@role_required('admin', 'director', 'secretary', 'teacher', 'parent')
def enrollment_detail(request, enrollment_id):
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'section__grade', 'academic_year'),
        id=enrollment_id
    )
    return render(request, 'enrollment/enrollment_detail.html', {'enrollment': enrollment})


@role_required('admin', 'director', 'secretary')
def enrollment_edit(request, enrollment_id):
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'section__grade', 'academic_year'),
        id=enrollment_id
    )

    if request.method == 'POST':
        form = EnrollmentForm(request.POST, instance=enrollment)
        if request.user.role == 'secretary' and not request.user.is_superuser:
            form.fields['student'].disabled = True

        if form.is_valid():
            updated = form.save(commit=False)
            # Secretaries can correct enrollment data but should not reassign a student to a different enrollment.
            if request.user.role == 'secretary' and not request.user.is_superuser:
                updated.student_id = enrollment.student_id
            updated.save()
            messages.success(request, "Matricula actualizada correctamente.")
            return redirect('enrollment_detail', enrollment_id=enrollment.id)
    else:
        form = EnrollmentForm(instance=enrollment)
        if request.user.role == 'secretary' and not request.user.is_superuser:
            form.fields['student'].disabled = True

    return render(request, 'enrollment/enrollment_edit_form.html', {
        'form': form,
        'enrollment': enrollment,
    })
